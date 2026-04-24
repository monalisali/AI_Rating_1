#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
重新执行优化提示词逻辑
从已有的输出Excel中读取AI回答，重新评分，如果平均分<70则优化提示词并重试
"""

import os
import sys
import json
import logging
from openpyxl import load_workbook
from datetime import datetime

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
logger = logging.getLogger(__name__)

# 导入app中的函数
from app import (
    load_system_prompt, score_answer, optimize_prompt,
    chat_with_confirmation, save_results_to_excel, save_results_to_html,
    _log_attempt_summary, load_saved_prompt
)

OUTPUT_DIR = 'outputs'


def read_results_from_excel(filepath):
    """从输出Excel中读取已有的问题和AI回答"""
    wb = load_workbook(filepath)
    ws = wb.active
    results = []
    row = 2
    while True:
        question = ws.cell(row=row, column=1).value
        if not question or str(question).strip() == '':
            break
        answer = ws.cell(row=row, column=2).value or ''
        reference_answer = ws.cell(row=row, column=3).value or ''
        results.append({
            'row': row,
            'question': str(question).strip(),
            'answer': str(answer).strip(),
            'reference_answer': str(reference_answer).strip() if reference_answer else ''
        })
        row += 1
    return results


def main():
    # 1. 找到最新的输出文件
    output_files = sorted(
        [f for f in os.listdir(OUTPUT_DIR) if f.endswith('.xlsx')],
        reverse=True
    )
    if not output_files:
        print("未找到输出文件")
        return

    latest_output = os.path.join(OUTPUT_DIR, output_files[0])
    logger.info(f"读取已有结果: {latest_output}")

    # 2. 读取已有结果
    results = read_results_from_excel(latest_output)
    if not results:
        logger.error("未读取到结果数据")
        return

    logger.info(f"共读取 {len(results)} 条结果")

    # 3. 重新评分（使用已有的AI回答）
    system_prompt = load_system_prompt()
    scoring_prompt_template = load_saved_prompt()

    logger.info("===== 重新评分（使用已有AI回答）=====")
    for r in results:
        if r['reference_answer']:
            scores = score_answer(r['question'], r['answer'], r['reference_answer'], scoring_prompt_template)
            r['scores'] = scores
            logger.info(f"  row={r['row']}: 总分={scores.get('total_score', '?')}")
        else:
            r['scores'] = None
            logger.info(f"  row={r['row']}: 无参考答案，跳过评分")

    # 4. 计算平均分
    avg_total = _log_attempt_summary(1, system_prompt, results)

    if avg_total >= 70:
        logger.info(f"平均总分 {avg_total:.1f} >= 70，无需优化")
        return

    # 5. 优化提示词
    logger.info(f"平均总分 {avg_total:.1f} < 70，开始优化提示词...")
    new_prompt = optimize_prompt(system_prompt, results)

    if new_prompt == system_prompt:
        logger.info("提示词未变更，终止")
        return

    logger.info("提示词已更新，用新提示词重新获取AI回答...")

    # 6. 用新提示词重新获取AI回答
    new_results = []
    for i, r in enumerate(results):
        logger.info(f"▶ 处理第{i+1}/{len(results)}题: {r['question'][:30]}...")
        try:
            new_answer = chat_with_confirmation(r['question'], system_prompt=new_prompt)
            scores = None
            if r['reference_answer']:
                scores = score_answer(r['question'], new_answer, r['reference_answer'], scoring_prompt_template)
            new_results.append({
                'row': r['row'],
                'question': r['question'],
                'answer': new_answer,
                'reference_answer': r['reference_answer'],
                'scores': scores,
                'success': True
            })
            if scores and scores.get('success'):
                logger.info(f"  总分={scores['total_score']}")
        except Exception as e:
            logger.error(f"  处理失败: {e}")
            new_results.append({
                'row': r['row'],
                'question': r['question'],
                'answer': f'处理失败: {str(e)}',
                'reference_answer': r['reference_answer'],
                'scores': None,
                'success': False
            })

    # 7. 记录第2次结果
    avg_total_2 = _log_attempt_summary(2, new_prompt, new_results)

    # 8. 对比
    logger.info("===== 优化前后对比 =====")
    logger.info(f"优化前平均总分: {avg_total:.1f}")
    logger.info(f"优化后平均总分: {avg_total_2:.1f}")

    # 9. 保存新结果
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename_xlsx = f"AI评估结果_优化后_{timestamp}.xlsx"
    output_filename_html = f"AI评估结果_优化后_{timestamp}.html"
    questions_with_answers = [
        (r['row'], r['question'], r['answer'], r.get('reference_answer', ''), r['scores'])
        for r in new_results
    ]
    save_results_to_excel(questions_with_answers, os.path.join(OUTPUT_DIR, output_filename_xlsx))
    save_results_to_html(questions_with_answers, os.path.join(OUTPUT_DIR, output_filename_html))
    logger.info(f"结果已保存: {output_filename_xlsx}, {output_filename_html}")


if __name__ == '__main__':
    main()
