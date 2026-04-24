#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Web应用 - 提供Excel上传页面并调用API处理
支持AI回答和语义对比打分
"""

import os
import re
import json
import ssl
import logging
import urllib.request
import queue
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side

# 日志配置：同时输出到控制台和文件
LOG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs')
os.makedirs(LOG_DIR, exist_ok=True)
LOG_FILE = os.path.join(LOG_DIR, f"optimize_{datetime.now().strftime('%Y%m%d')}.log")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

_log_fmt = logging.Formatter('%(asctime)s %(levelname)s %(message)s')

_file_handler = logging.FileHandler(LOG_FILE, encoding='utf-8')
_file_handler.setLevel(logging.INFO)
_file_handler.setFormatter(_log_fmt)
logger.addHandler(_file_handler)

_console_handler = logging.StreamHandler()
_console_handler.setLevel(logging.INFO)
_console_handler.setFormatter(_log_fmt)
logger.addHandler(_console_handler)

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 最大16MB
app.config['JSON_AS_ASCII'] = False

# 确保文件夹存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# Prompt.md 路径
PROMPT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Prompt.md')


def load_system_prompt():
    """加载 Prompt.md 作为知识库API的自定义系统提示词"""
    if os.path.exists(PROMPT_FILE):
        with open(PROMPT_FILE, 'r', encoding='utf-8') as f:
            return f.read().strip()
    logger.warning(f"Prompt.md 不存在: {PROMPT_FILE}")
    return ""


# 评分模型配置 - 使用内网API
SCORING_API_URL = os.environ.get('ANTHROPIC_BASE_URL', 'http://ai.tech.tax.asia.pwcinternal.com:3002') + '/v1/chat/completions'
SCORING_API_KEY = os.environ.get('ANTHROPIC_AUTH_TOKEN', '')
SCORING_MODEL = os.environ.get('ANTHROPIC_MODEL', 'glm-coding-5-8')


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

"""
model参数
百炼模型：glm-5.1
GTS模型：saas.glm-5.1
"""

def request_api(message: str, session_id: str = "", custom_system_prompt: str = "") -> tuple:
    """请求知识库API接口"""
    url = 'https://ai.tech.tax.asia.pwcinternal.com:5007/api/chat-stream'
    payload = {
        'message': message,
        'session_id': session_id,
        'model': 'glm-5.1'
    }
    if custom_system_prompt:
        payload['custom_system_prompt'] = custom_system_prompt
    data = json.dumps(payload).encode('utf-8')

    req = urllib.request.Request(url, data=data, headers={
        'Content-Type': 'application/json',
        'Accept': 'text/event-stream'
    })

    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    with urllib.request.urlopen(req, context=ctx, timeout=300) as response:
        returned_session_id = response.headers.get('X-Session-Id', '')
        chunks = []
        while True:
            chunk = response.read(8192)
            if not chunk:
                break
            chunks.append(chunk)
        result = b''.join(chunks).decode('utf-8')

    return result, returned_session_id


def parse_response(api_response: str) -> dict:
    """解析SSE流式API响应"""
    contents = []
    for line in api_response.strip().split('\n'):
        line = line.strip()
        if line.startswith('data:'):
            try:
                data = json.loads(line[5:].strip())
                if data.get('type') == 'content':
                    contents.append(data.get('content', ''))
                elif data.get('type') == 'error':
                    logger.error(f"[parse_response] API返回错误: {data}")
            except json.JSONDecodeError:
                pass

    full = ''.join(contents)

    # 截取"核心发现"之后的内容
    marker = '核心发现'
    idx = full.find(marker)
    if idx != -1:
        full = full[idx:]

    return {'full_content': full}


def is_confirmation_question(content: str) -> bool:
    """判断内容是否是确认问句"""
    keywords = [
        '请问以上关键词是否需要调整或补充', '确认后我将开始',
        '请问这样理解是否正确', '您是否有需要补充或调整',
        '是否需要调整', '请确认', '确认后', '以上关键词是否准确',
        '是否需要添加', '请问以上', '确认后开始', '是否准确',
        '请告知', '是否继续', '我将开始', '以上内容是否',
        '是否合适', '希望调整', '是否需要修改', '是否满意',
        '是否同意', '请告诉我您的修改', '需要调整请告诉我',
    ]
    return any(kw in content for kw in keywords)


def is_incomplete_answer(content: str) -> bool:
    """判断内容是否是不完整的中间步骤（没有核心发现且较短）"""
    if '核心发现' in content:
        return False
    # 内容过短说明还没拿到完整答案
    return len(content.strip()) < 200


def chat_with_confirmation(question: str, max_rounds: int = 8, system_prompt: str = "") -> str:
    """执行多轮对话，自动处理确认和等待完整答案"""
    session_id = ""
    current_message = question

    for _ in range(max_rounds):
        api_response, session_id = request_api(current_message, session_id, custom_system_prompt=system_prompt)
        content = parse_response(api_response)['full_content']
        if is_confirmation_question(content):
            current_message = "同意，请使用这些关键词进行搜索，不需要调整。"
        elif is_incomplete_answer(content):
            current_message = "继续"
        else:
            return content

    return content


def request_scoring_api(prompt: str) -> str:
    """请求内网AI评分API"""
    data = json.dumps({
        'model': SCORING_MODEL,
        'messages': [{'role': 'user', 'content': prompt}]
    }).encode('utf-8')

    req = urllib.request.Request(SCORING_API_URL, data=data, headers={
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {SCORING_API_KEY}'
    })

    with urllib.request.urlopen(req, timeout=300) as response:
        result = json.loads(response.read().decode('utf-8'))
        if 'choices' not in result or not result['choices']:
            raise ValueError(f"API返回格式异常: {json.dumps(result, ensure_ascii=False)[:300]}")
        return result['choices'][0]['message']['content']


# 默认评分提示词模板（可通过前端覆盖）
DEFAULT_SCORING_PROMPT = """请作为一名专业的税务领域评估专家，对以下AI回答进行评分。

【问题】
{question}

【参考答案】
{reference_answer}

【AI回答】
{ai_answer}

请从以下3个维度进行评分，并以JSON格式返回结果：

1. 答案准确性（60分）：
   - 将AI回答中"核心发现"部分的每一条要点与参考答案进行逐条语义对比，重点判断核心结论是否准确、有无遗漏或偏差。"核心发现"是回答的精华摘要，其准确性权重最高。
   - 对回答整体内容与参考答案进行全面语义对比，评估事实、数据、税率、政策适用等方面是否准确。
   - 评估回答是否针对问题所涉及的具体地区（如中国大陆、香港、特定省市等）的税收政策和规定进行了准确回答。如果问题涉及特定地区，回答是否正确区分了该地区的特殊规定；如果回答了不相关的地区政策应扣分。若问题不涉及地区性差异则不扣分。
   - 重点检查"核心发现"部分的结论与回答后续展开的详细分析、案例、数据之间是否存在矛盾或冲突。例如：核心发现说适用税率A，但详细分析中又说适用税率B；或核心发现说某政策适用，但详细分析中列出了不满足的条件。存在此类矛盾应扣分。

   【严重偏差判定——符合以下任一情形，准确性直接打0分】：
   a) 遗漏关键税务处理步骤：参考答案中的核心税务调整步骤（如视同销售后的费用扣除基数联动调整、特定申报表行次填报、跨表勾稽关系等）在AI回答中完全未提及，导致整条计算链条断裂。
   b) 核心计算基数错误：AI回答中用于关键计算的基础数据（如收入额、扣除限额的计税基础、视同销售的公允价值等）与参考答案不一致，导致后续计算结果全部偏误。
   c) 引入题干明确排除的无关内容且挤占关键篇幅：题干已限定范围（如"仅企业所得税""不考虑增值税"），AI回答却大篇幅讨论题干明确排除的税种或无关政策，且该无关内容导致关键要点被遗漏或论述不足。
   d) 核心要点遗漏过半：参考答案包含多个独立的核心结论或处理步骤，AI回答遗漏其中超过半数。
   - 以上方面均准确无误、核心发现与参考答案完全一致且无矛盾给60分。
   - 当准确性评分为0分时，法条援引度和总结完整度也全部直接置为0分。

2. 法条援引度（20分）：
   - 检查AI回答中引用的法规条款是否正确、是否与参考答案中提及的法条一致。
   - 检查回答末尾或正文中的引用法规列表/参考资料，判断这些法规是否在回答正文中被实际引用和讨论。
   - 检查回答正文中提到的法规是否都在引用列表中列出，确保引用列表与正文引用完全对应。
   - 引用正确完整、列表与正文完全对应给20分。

3. 总结完整度（20分）：
   - 评估"核心发现"作为总结是否完整，是否涵盖了问题的所有关键方面，是否遗漏了参考答案中的重要要点。
   - 检查回答是否对问题中隐含的子问题或特殊情况都给出了回应。
   - 总结全面、覆盖完整给20分。

请严格按照以下JSON格式返回（不要添加任何其他文字）：
{{
    "accuracy_score": 分数,
    "accuracy_reason": "评分说明",
    "citation_score": 分数,
    "citation_reason": "评分说明",
    "summary_score": 分数,
    "summary_reason": "评分说明"
}}"""


def optimize_prompt(current_prompt: str, results_with_scores: list, attempt: int = 1) -> str:
    """
    根据评分结果，定向优化系统提示词

    Args:
        current_prompt: 当前使用的系统提示词
        results_with_scores: 列表，每项包含 question, answer, reference_answer, scores

    Returns:
        优化后的新系统提示词
    """
    logger.info("[optimize_prompt] 开始优化提示词...")

    # 构建评分详情（不截取）
    details = []
    low_dims = {'accuracy': False, 'citation': False, 'summary': False}
    for i, r in enumerate(results_with_scores):
        scores = r.get('scores', {})
        detail = f"问题{i+1}: {r['question']}\n"
        detail += f"AI回答: {r['answer']}\n"
        if r.get('reference_answer'):
            detail += f"参考答案: {r['reference_answer']}\n"
        if scores and scores.get('success'):
            detail += f"总分: {scores['total_score']}/100\n"
            detail += f"准确性: {scores['accuracy_score']}/60 - {scores['accuracy_reason']}\n"
            detail += f"法条援引: {scores['citation_score']}/20 - {scores['citation_reason']}\n"
            detail += f"总结完整度: {scores['summary_score']}/20 - {scores['summary_reason']}\n"
            if scores['accuracy_score'] < 48:
                low_dims['accuracy'] = True
            if scores['citation_score'] < 16:
                low_dims['citation'] = True
            if scores['summary_score'] < 16:
                low_dims['summary'] = True
        else:
            detail += "评分失败\n"
        details.append(detail)

    # 根据低分维度生成定向优化指令
    focus_parts = []
    if low_dims['accuracy']:
        focus_parts.append(
            "- 【准确性偏低】重点优化搜索策略（步骤4和4.6），确保AI搜索更全面、不遗漏关键法规，"
            "加强法规内容与问题的关联分析，确保核心结论准确。不要修改与搜索无关的部分。"
        )
    if low_dims['citation']:
        focus_parts.append(
            "- 【法条援引偏低】重点优化引用列表与校验部分（步骤6），确保AI在回答中实际引用和讨论每条法规，"
            "引用列表与正文引用完全对应。不要修改与引用无关的部分。"
        )
    if low_dims['summary']:
        focus_parts.append(
            "- 【总结完整度偏低】重点优化核心发现和总结的生成要求（步骤5），确保AI的总结覆盖问题的所有关键方面，"
            "不遗漏重要要点。不要修改与总结无关的部分。"
        )

    focus_text = '\n'.join(focus_parts) if focus_parts else "- 整体评分偏低，请全面检查并优化系统提示词中的薄弱环节。"

    optimize_instruction = f"""你是一个专业的提示词优化专家。当前有一个用于税务法规知识库AI助手的系统提示词，但使用该提示词后AI回答的评分不够理想。

请根据以下评分数据中体现的具体薄弱维度，**只针对低分维度相关部分进行定向修改**，其他部分必须原样保留不得改动。

【当前系统提示词】
{current_prompt}

【各题评分详情】
{chr(10).join(details)}

【需要优化的维度】
{focus_text}

【优化规则】
1. 只修改与低分维度直接相关的步骤或段落，其他所有内容必须原样保留
2. 不要重构、不要重写、不要删除已有的有效规则
3. 修改时要具体、可操作，添加明确的指令而非模糊的建议
4. 输出完整的系统提示词（包含未修改的部分），不要省略任何部分
5. 不要输出任何解释说明，只输出新的系统提示词本身"""

    try:
        new_prompt = request_scoring_api(optimize_instruction)
        logger.info(f"[optimize_prompt] 优化完成，原始长度={len(current_prompt)}, 新长度={len(new_prompt)}")
        # 追加保存到当次运行的优化提示词文件
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        prompt_file = os.path.join(LOG_DIR, f"optimized_prompt_{ts[:8]}.txt")
        with open(prompt_file, 'a', encoding='utf-8') as f:
            f.write(f"\n{'='*60}\n")
            f.write(f"===== 第{attempt}次优化生成的提示词 ({ts}) =====\n")
            f.write(f"{'='*60}\n\n")
            f.write(new_prompt)
            f.write("\n")
        logger.info(f"[optimize_prompt] 优化后提示词已追加到: {prompt_file}")
        return new_prompt
    except Exception as e:
        logger.error(f"提示词优化失败: {e}")
        return current_prompt


def score_answer(question: str, ai_answer: str, reference_answer: str, scoring_prompt_template: str = None) -> dict:
    """使用AI对回答进行语义对比打分"""
    # 使用自定义提示词模板或默认模板
    template = scoring_prompt_template or DEFAULT_SCORING_PROMPT
    scoring_prompt = template.format(
        question=question,
        reference_answer=reference_answer,
        ai_answer=ai_answer
    )

    failed = {
        'success': False,
        'accuracy_score': 0, 'accuracy_reason': '评分失败',
        'citation_score': 0, 'citation_reason': '评分失败',
        'summary_score': 0, 'summary_reason': '评分失败',
        'total_score': 0
    }

    last_error = None
    for attempt in range(3):
        try:
            content = request_scoring_api(scoring_prompt)

            json_start = content.find('{')
            json_end = content.rfind('}') + 1
            if json_start == -1 or json_end <= json_start:
                last_error = f"评分返回无JSON: {content[:200]}"
                logger.error(f"第{attempt+1}次尝试 - {last_error}")
                continue

            scores = json.loads(content[json_start:json_end])
            score_keys = ['accuracy_score', 'citation_score', 'summary_score']
            result = {
                'success': True,
                **{k: scores.get(k, 0) for k in score_keys},
                **{k.replace('score', 'reason'): scores.get(k.replace('score', 'reason'), '') for k in score_keys},
                'total_score': sum(scores.get(k, 0) for k in score_keys)
            }
            logger.info(f"[score_answer] 评分完成 → 总分={result['total_score']}, 准确性={result['accuracy_score']}, 法条={result['citation_score']}, 总结={result['summary_score']}")
            return result
        except Exception as e:
            last_error = str(e)
            logger.error(f"第{attempt+1}次评分失败: {e}")
            if attempt < 2:
                import time
                time.sleep(2)

    # 3次都失败，把具体错误写入reason
    failed['accuracy_reason'] = f'评分失败({last_error})' if last_error else '评分失败'
    failed['citation_reason'] = failed['accuracy_reason']
    failed['summary_reason'] = failed['accuracy_reason']
    return failed


def read_questions_from_excel(filepath):
    """从Excel文件读取问题(B列)、现有AI回答(C列)和建议答案(E列)"""
    wb = load_workbook(filepath)
    ws = wb.active
    questions = []
    row = 2

    while True:
        cell_value = ws[f'B{row}'].value
        if cell_value is None or str(cell_value).strip() == '':
            break

        existing_answer = ws[f'C{row}'].value
        existing_answer = str(existing_answer).strip() if existing_answer else ''

        reference_answer = ws[f'E{row}'].value
        reference_answer = str(reference_answer).strip() if reference_answer else ''
        questions.append((row, str(cell_value).strip(), existing_answer, reference_answer))
        row += 1

    return questions


def save_results_to_excel(questions_with_answers, output_filepath):
    """保存结果到新Excel，只保留问题、建议答案、评分情况"""
    wb = Workbook()
    ws = wb.active
    ws.title = '评估结果'

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        ('问题', 50),
        ('AI回答', 100),
        ('建议答案', 80),
        ('答案准确性(60分)', 12), ('答案准确性说明', 40),
        ('法条援引度(20分)', 12), ('法条援引度说明', 40),
        ('总结完整度(20分)', 12), ('总结完整度说明', 40),
        ('总分(100分)', 10)
    ]

    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, (row_num, question, answer, reference_answer, scores) in enumerate(questions_with_answers, 2):
        # 问题
        cell = ws.cell(row=row_idx, column=1, value=question)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border

        # AI回答
        cell = ws.cell(row=row_idx, column=2, value=answer)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border

        # 建议答案
        cell = ws.cell(row=row_idx, column=3, value=reference_answer or '')
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border

        if scores:
            score_data = [
                scores.get('accuracy_score', 0),
                scores.get('accuracy_reason', ''),
                scores.get('citation_score', 0),
                scores.get('citation_reason', ''),
                scores.get('summary_score', 0),
                scores.get('summary_reason', ''),
                scores.get('total_score', 0)
            ]
            for i, value in enumerate(score_data):
                cell = ws.cell(row=row_idx, column=4 + i, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = thin_border

    wb.save(output_filepath)


def save_results_to_html(questions_with_answers, output_filepath):
    """保存结果到HTML文件，PwC风格"""
    rows_html = ''
    for row_idx, (row_num, question, answer, reference_answer, scores) in enumerate(questions_with_answers, 1):
        bg = '#FFFFFF' if row_idx % 2 == 1 else '#F4F4F4'
        scores_html = ''
        if scores and scores.get('success'):
            score_items = [
                ('答案准确性', scores.get('accuracy_score', 0), scores.get('accuracy_reason', '')),
                ('法条援引度', scores.get('citation_score', 0), scores.get('citation_reason', '')),
                ('总结完整度', scores.get('summary_score', 0), scores.get('summary_reason', '')),
            ]
            items_html = ''
            for label, val, reason in score_items:
                color = '#D04A02' if val < 15 else '#2A9D8F'
                items_html += f'''<div style="display:inline-block;width:19%;text-align:center;padding:6px 0;">
                    <div style="font-size:11px;color:#8A8C8E;">{label}</div>
                    <div style="font-size:20px;font-weight:700;color:{color};">{val}/20</div>
                    <div style="font-size:11px;color:#8A8C8E;text-align:left;line-height:1.3;">{reason}</div>
                </div>'''
            total = scores.get('total_score', 0)
            total_color = '#D04A02' if total < 75 else '#2A9D8F'
            items_html += f'''<div style="display:inline-block;width:19%;text-align:center;padding:6px 0;background:#2D2D2D;border-radius:3px;color:white;">
                <div style="font-size:11px;color:rgba(255,255,255,0.7);">总分</div>
                <div style="font-size:20px;font-weight:700;color:{total_color};">{total}/100</div>
            </div>'''
            scores_html = f'<div style="margin-top:8px;padding:8px;background:white;border-radius:3px;border:1px solid #E8E8E8;">{items_html}</div>'
        elif scores:
            scores_html = f'<div style="margin-top:8px;padding:8px;background:#FEF2F2;border-radius:3px;color:#CB333B;font-size:13px;">评分失败</div>'

        safe_q = question.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        safe_ans = answer.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        safe_ref = (reference_answer or '').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

        rows_html += f'''<tr style="background:{bg};vertical-align:top;">
            <td style="padding:10px;border:1px solid #E8E8E8;width:12%;">{safe_q}</td>
            <td style="padding:10px;border:1px solid #E8E8E8;width:30%;">{safe_ans}</td>
            <td style="padding:10px;border:1px solid #E8E8E8;width:18%;">{safe_ref}</td>
            <td style="padding:10px;border:1px solid #E8E8E8;width:40%;">{scores_html}</td>
        </tr>'''

    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>AI回答评估结果</title>
<style>
body {{ font-family: Helvetica, Arial, 'PingFang SC', 'Microsoft YaHei', sans-serif; background:#F4F4F4; padding:20px; color:#2D2D2D; }}
h1 {{ font-size:20px; font-weight:600; border-bottom:2px solid #D04A02; padding-bottom:10px; }}
</style>
</head>
<body>
<h1>AI回答评估结果</h1>
<table style="width:100%;border-collapse:collapse;font-size:13px;">
<thead><tr style="background:#2D2D2D;color:white;">
<th style="padding:10px;border:1px solid #2D2D2D;width:12%;text-align:left;">问题</th>
<th style="padding:10px;border:1px solid #2D2D2D;width:30%;text-align:left;">AI回答</th>
<th style="padding:10px;border:1px solid #2D2D2D;width:18%;text-align:left;">建议答案</th>
<th style="padding:10px;border:1px solid #2D2D2D;width:40%;text-align:left;">评分情况</th>
</tr></thead>
<tbody>{rows_html}</tbody>
</table>
</body></html>'''

    with open(output_filepath, 'w', encoding='utf-8') as f:
        f.write(html)


@app.route('/')
def index():
    return render_template('index.html')


# 提示词持久化文件路径
SCORING_PROMPT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'scoring_prompt.txt')


def load_saved_prompt():
    """加载已保存的评分提示词，如果没有则返回默认模板"""
    if os.path.exists(SCORING_PROMPT_FILE):
        try:
            with open(SCORING_PROMPT_FILE, 'r', encoding='utf-8') as f:
                saved = f.read().strip()
                if saved:
                    return saved
        except Exception as e:
            logger.warning(f"读取保存的提示词失败: {e}")
    return DEFAULT_SCORING_PROMPT


def save_prompt(prompt_text):
    """保存评分提示词到文件"""
    with open(SCORING_PROMPT_FILE, 'w', encoding='utf-8') as f:
        f.write(prompt_text)


@app.route('/default_prompt', methods=['GET'])
def get_default_prompt():
    """返回默认评分提示词模板"""
    return jsonify({'prompt': DEFAULT_SCORING_PROMPT})


@app.route('/saved_prompt', methods=['GET'])
def get_saved_prompt():
    """返回已保存的评分提示词（优先返回用户自定义，没有则返回默认）"""
    return jsonify({'prompt': load_saved_prompt()})


@app.route('/save_prompt', methods=['POST'])
def save_prompt_api():
    """保存用户编辑的评分提示词"""
    data = request.json
    prompt_text = data.get('prompt', '').strip()
    if not prompt_text:
        return jsonify({'error': '提示词不能为空'}), 400
    try:
        save_prompt(prompt_text)
        return jsonify({'success': True, 'message': '提示词已保存'})
    except Exception as e:
        return jsonify({'error': f'保存失败: {str(e)}'}), 500


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': '没有上传文件'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400

    if not (file and allowed_file(file.filename)):
        return jsonify({'error': '不支持的文件格式'}), 400

    # 保留中文文件名，只去掉不安全字符
    safe_name = re.sub(r'[\\/*?:"<>|]', '_', file.filename)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{timestamp}_{safe_name}"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    try:
        questions = read_questions_from_excel(filepath)
    except Exception as e:
        return jsonify({'error': f'读取Excel文件失败: {str(e)}'}), 400

    return jsonify({
        'success': True,
        'filename': filename,
        'question_count': len(questions),
        'questions': [{'row': r, 'question': q, 'has_existing_answer': bool(ea), 'has_reference': bool(ref)} for r, q, ea, ref in questions]
    })


def process_single_question(row_num, question, reference_answer, enable_scoring, scoring_prompt_template=None, system_prompt="", existing_answer=""):
    """处理单个问题：获取AI回答 + 评分（如果有现有AI回答则跳过API调用）"""
    try:
        if existing_answer:
            answer = existing_answer
        else:
            answer = chat_with_confirmation(question, system_prompt=system_prompt)

        scores = None
        if enable_scoring and reference_answer:
            logger.info(f"[process] row={row_num} 开始评分...")
            scores = score_answer(question, answer, reference_answer, scoring_prompt_template)
            logger.info(f"[process] row={row_num} 评分完成 → 总分={scores.get('total_score', '?')}")

        return {
            'row': row_num,
            'question': question,
            'answer': answer,
            'reference_answer': reference_answer,
            'scores': scores,
            'success': True
        }
    except Exception as e:
        logger.error(f"问题处理失败 (行{row_num}): {e}")
        return {
            'row': row_num,
            'question': question,
            'answer': f'处理失败: {str(e)}',
            'reference_answer': reference_answer,
            'scores': None,
            'success': False
        }


def _log_attempt_summary(attempt, system_prompt, results):
    """记录单次尝试的日志：提示词摘要 + 每题评分 + 整体统计，返回平均总分"""
    logger.info(f"===== 第{attempt}次尝试 =====")
    logger.info(f"系统提示词(前200字): {system_prompt[:200]}")

    scored = [r for r in results if r.get('scores', {}).get('success')]
    for r in results:
        scores = r.get('scores', {})
        if scores and scores.get('success'):
            logger.info(
                f"  题目(row={r['row']}): 总分={scores['total_score']}, "
                f"准确性={scores['accuracy_score']}/60, "
                f"法条援引={scores['citation_score']}/20, "
                f"总结完整度={scores['summary_score']}/20"
            )
        else:
            logger.info(f"  题目(row={r['row']}): 评分失败")

    if scored:
        n = len(scored)
        avg_total = sum(r['scores']['total_score'] for r in scored) / n
        avg_acc = sum(r['scores']['accuracy_score'] for r in scored) / n
        avg_cit = sum(r['scores']['citation_score'] for r in scored) / n
        avg_sum = sum(r['scores']['summary_score'] for r in scored) / n
        logger.info(
            f"整体统计: 平均总分={avg_total:.1f}, "
            f"平均准确性={avg_acc:.1f}, "
            f"平均法条援引={avg_cit:.1f}, "
            f"平均总结完整度={avg_sum:.1f}"
        )
        return avg_total
    return 0


@app.route('/process', methods=['POST'])
def process_questions():
    """SSE流式处理接口，支持自动优化提示词重试"""
    data = request.json
    filename = data.get('filename')
    enable_scoring = data.get('enable_scoring', False)
    scoring_prompt_template = data.get('scoring_prompt', None)
    thread_count = data.get('thread_count', 2)
    thread_count = max(1, min(8, int(thread_count)))
    max_attempts = max(1, min(20, int(data.get('max_attempts', 20))))

    def sse_error(msg):
        def gen():
            yield f"data: {json.dumps({'type': 'error', 'message': msg}, ensure_ascii=False)}\n\n"
        return app.response_class(gen(), mimetype='text/event-stream',
                                  headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})

    if not filename:
        return sse_error('缺少文件名')

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(filepath):
        return sse_error('文件不存在')

    questions = read_questions_from_excel(filepath)
    total = len(questions)
    system_prompt = load_system_prompt()
    logger.info(f"========== 开始处理 {total}题 | {thread_count}线程 | 评分{'开' if enable_scoring else '关'} | 最多{max_attempts}次尝试 ==========")
    has_existing = any(ea for _, _, ea, _ in questions)

    def generate():
        nonlocal system_prompt

        all_attempt_logs = []
        results = []

        for attempt in range(1, max_attempts + 1):
            try:
                if attempt == 1:
                    logger.info(f"===== 第{attempt}次尝试（使用{'现有AI回答' if has_existing else 'API调用'}）=====")
                else:
                    logger.info(f"===== 第{attempt}次尝试（使用优化后提示词，重新调API）=====")

                result_queue = queue.Queue()

                # 第1次尝试：如果有现有AI回答则使用（跳过API调用）；第2次及以后：用新提示词重新调API
                is_first_attempt_with_existing = (attempt == 1 and has_existing)

                def worker(row_num, question, existing_answer, reference_answer):
                    ea = existing_answer if is_first_attempt_with_existing else ""
                    result = process_single_question(
                        row_num, question, reference_answer,
                        enable_scoring, scoring_prompt_template, system_prompt, ea
                    )
                    result_queue.put(result)

                # 提交所有任务
                with ThreadPoolExecutor(max_workers=thread_count) as executor:
                    futures = []
                    for row_num, question, existing_answer, reference_answer in questions:
                        futures.append(executor.submit(worker, row_num, question, existing_answer, reference_answer))

                    completed = 0
                    results = []
                    while completed < total:
                        result = result_queue.get()
                        results.append(result)
                        completed += 1

                        event = {
                            'type': 'progress',
                            'current': completed,
                            'total': total,
                            'percentage': int(completed / total * 100),
                            'attempt': attempt,
                            'max_attempts': max_attempts,
                            'result': result
                        }
                        yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"

                results.sort(key=lambda r: r['row'])

                # 记录本次尝试的评分摘要
                avg_total = _log_attempt_summary(attempt, system_prompt, results)
                all_attempt_logs.append({
                    'attempt': attempt,
                    'system_prompt': system_prompt,
                    'avg_total': avg_total,
                    'results': results
                })

                # 判断是否需要重试
                need_retry = (
                    enable_scoring
                    and attempt < max_attempts
                    and avg_total < 70
                    and avg_total >= 0
                    and len([r for r in results if r.get('scores', {}).get('success')]) > 0
                )

                if need_retry:
                    logger.info(f">>> 平均总分 {avg_total:.1f} < 70，触发第{attempt}次提示词优化...")
                    yield f"data: {json.dumps({'type': 'optimizing', 'attempt': attempt, 'avg_score': round(avg_total, 1)}, ensure_ascii=False)}\n\n"

                    old_prompt = system_prompt
                    system_prompt = optimize_prompt(system_prompt, results, attempt)

                    if system_prompt != old_prompt:
                        logger.info(f">>> 提示词已优化，将用新提示词重新获取AI回答并评分")
                    else:
                        logger.info(">>> 提示词未变更，终止重试")
                        break
                else:
                    break

            except Exception as e:
                logger.error(f"第{attempt}次尝试异常: {e}")
                yield f"data: {json.dumps({'type': 'error', 'message': f'第{attempt}次尝试异常: {str(e)}'}, ensure_ascii=False)}\n\n"
                if not results:
                    return

        # 按行号排序后保存最终结果
        try:
            results.sort(key=lambda r: r['row'])
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            output_filename_xlsx = f"AI评估结果_{timestamp}.xlsx"
            output_filename_html = f"AI评估结果_{timestamp}.html"
            output_xlsx = os.path.join(app.config['OUTPUT_FOLDER'], output_filename_xlsx)
            output_html = os.path.join(app.config['OUTPUT_FOLDER'], output_filename_html)
            questions_with_answers = [(r['row'], r['question'], r['answer'], r.get('reference_answer', ''), r['scores']) for r in results]
            logger.info(f"保存结果 → Excel={output_filename_xlsx}, HTML={output_filename_html}")
            save_results_to_excel(questions_with_answers, output_xlsx)
            save_results_to_html(questions_with_answers, output_html)
        except Exception as e:
            logger.error(f"保存结果异常: {e}")
            output_filename_xlsx = ""
            output_filename_html = ""

        # 最终汇总
        try:
            logger.info("=" * 60)
            if len(all_attempt_logs) > 1:
                first_avg = all_attempt_logs[0]['avg_total']
                best = max(all_attempt_logs, key=lambda x: x['avg_total'])
                logger.info(f"各次尝试平均总分: " + " → ".join(f"第{e['attempt']}次={e['avg_total']:.1f}" for e in all_attempt_logs))
                logger.info(f"评分提升: {first_avg:.1f} → {best['avg_total']:.1f} (提高了 {best['avg_total'] - first_avg:.1f} 分)")
                logger.info(f"===== 最高评分提示词(第{best['attempt']}次, 平均总分={best['avg_total']:.1f})完整内容 =====\n"
                            f"{best['system_prompt']}\n"
                            f"===== 提示词结束 =====")
            logger.info(f"========== 全部完成，共{attempt}次尝试，结果已保存 ==========")
        except Exception as e:
            logger.error(f"汇总日志异常: {e}")

        yield f"data: {json.dumps({'type': 'complete', 'output_filename': output_filename_xlsx, 'output_filename_html': output_filename_html, 'attempts': attempt}, ensure_ascii=False)}\n\n"

    return app.response_class(generate(), mimetype='text/event-stream',
                              headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


@app.route('/api/evaluate', methods=['POST'])
def api_evaluate():
    """
    外部API接口 - 批量评估问题
    入参JSON格式：
    {
        "items": [
            {"question": "问题1", "reference_answer": "建议答案1"},
            {"question": "问题2", "reference_answer": "建议答案2"}
        ],
        "scoring_prompt": "可选，自定义评分提示词模板",
        "thread_count": 4
    }
    返回JSON格式：
    {
        "success": true,
        "count": 2,
        "results": [
            {
                "question": "问题1",
                "answer": "AI回答内容",
                "reference_answer": "建议答案1",
                "scores": { ... }
            }
        ]
    }
    """
    data = request.json
    if not data or 'items' not in data:
        return jsonify({'success': False, 'error': '请求体需包含 items 数组'}), 400

    items = data.get('items', [])
    if not items or not isinstance(items, list):
        return jsonify({'success': False, 'error': 'items 不能为空'}), 400

    # 校验每条数据
    for i, item in enumerate(items):
        if not item.get('question'):
            return jsonify({'success': False, 'error': f'第{i+1}条数据缺少 question 字段'}), 400

    scoring_prompt_template = data.get('scoring_prompt') or load_saved_prompt()
    thread_count = max(1, min(8, int(data.get('thread_count', 4))))
    max_attempts = max(1, min(20, int(data.get('max_attempts', 20))))
    system_prompt = load_system_prompt()

    def process_one(idx, item):
        question = item['question']
        reference_answer = item.get('reference_answer', '')
        try:
            answer = chat_with_confirmation(question, system_prompt=system_prompt)
            scores = None
            if reference_answer:
                scores = score_answer(question, answer, reference_answer, scoring_prompt_template)
            return {
                'index': idx,
                'question': question,
                'answer': answer,
                'reference_answer': reference_answer,
                'scores': scores,
                'success': True
            }
        except Exception as e:
            logger.error(f"API评估失败 (第{idx+1}条): {e}")
            return {
                'index': idx,
                'question': question,
                'answer': f'处理失败: {str(e)}',
                'reference_answer': reference_answer,
                'scores': None,
                'success': False
            }

    all_attempt_logs = []
    logger.info(f"[api_evaluate] 开始: {len(items)}题, max_attempts={max_attempts}, threads={thread_count}")

    for attempt in range(1, max_attempts + 1):
        logger.info(f"[api_evaluate] ===== 第{attempt}/{max_attempts}次尝试 =====")
        results = [None] * len(items)
        with ThreadPoolExecutor(max_workers=thread_count) as executor:
            futures = {executor.submit(process_one, i, item): i for i, item in enumerate(items)}
            for future in futures:
                result = future.result()
                results[result['index']] = result

        avg_total = _log_attempt_summary(attempt, system_prompt, results)
        all_attempt_logs.append({
            'attempt': attempt,
            'system_prompt': system_prompt,
            'avg_total': avg_total
        })

        need_retry = attempt < max_attempts and avg_total < 70 and len([r for r in results if r.get('scores', {}).get('success')]) > 0
        if need_retry:
            logger.info(f"api_evaluate: 平均总分 {avg_total:.1f} < 70，优化提示词...")
            old_prompt = system_prompt
            system_prompt = optimize_prompt(system_prompt, results)
            if system_prompt == old_prompt:
                logger.info("提示词未变更，终止重试")
                break
        else:
            break

    # 记录各次尝试对比
    if len(all_attempt_logs) > 1:
        logger.info("===== api_evaluate 各次尝试评分对比 =====")
        for log_entry in all_attempt_logs:
            logger.info(
                f"第{log_entry['attempt']}次: 平均总分={log_entry['avg_total']:.1f}, "
                f"提示词前200字={log_entry['system_prompt'][:200]}"
            )

        best = max(all_attempt_logs, key=lambda x: x['avg_total'])
        logger.info(
            f"===== 最高评分提示词(第{best['attempt']}次, 平均总分={best['avg_total']:.1f})完整内容 =====\n"
            f"{best['system_prompt']}\n"
            f"===== 提示词结束 ====="
        )

    logger.info(f"[api_evaluate] 全部完成，共尝试{attempt}次")

    return jsonify({
        'success': True,
        'count': len(results),
        'results': results,
        'attempts': attempt
    })


@app.route('/download/<filename>')
def download_file(filename):
    filepath = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return jsonify({'error': '文件不存在'}), 404


if __name__ == '__main__':
    app.run(debug=True, port=5001, host='0.0.0.0')
