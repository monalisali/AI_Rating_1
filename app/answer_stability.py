#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
回答稳定性模块 - 四个Skill并行处理 + 动态提示词组装
解决同一问题+所有文章下AI回答不稳定的问题
"""

import os
import json
import queue
import time
import requests
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor

from flask import Blueprint, request as flask_request, jsonify, Response
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side

from app.app import (
    get_config, logger,
    ROOT_DIR, allowed_file, request_api, parse_response,
    is_confirmation_question, is_incomplete_answer
)

stability_bp = Blueprint('answer_stability', __name__)

UPLOAD_FOLDER = os.path.join(ROOT_DIR, 'uploads')
OUTPUT_FOLDER = os.path.join(ROOT_DIR, 'outputs')

# ===================== 后处理配置（三层过滤 按计算后的阈值过滤） =====================
# 是否启用三层过滤（相关性→置信度→冲突消解），False时跳过过滤，"过滤后"栏位填"阈值过滤逻辑未启用"
FILTER_ENABLED = False

# 相关性过滤阈值（BGE-M3余弦相似度，产物文本与用户问题的最低相似度）
FILTER_REL_CC = 0.60          # 条件-结论对
FILTER_REL_SCENE = 0.50       # 政策场景标签
FILTER_REL_ASSERTION = 0.40   # 概念关系断言
FILTER_REL_TIME = 0.50        # 时间约束

# 置信度过滤阈值（基于source_count和证据质量计算的最低置信度）
FILTER_CONF_CC = 0.60         # 条件-结论对
FILTER_CONF_ASSERTION = 0.60  # 概念关系断言
FILTER_CONF_TIME = 0.80       # 时间约束（时间信息通常可靠，提高阈值）
# 政策场景不设置信度过滤，仅依赖相关性

# 专门性关键词（冲突消解中判断断言是否针对特定主体）
SPECIFIC_KEYWORDS = [
    "西部大开发", "高新技术企业", "小型微利", "软件企业",
    "集成电路", "经济特区", "浦东新区", "海南自贸港"
]

# ===================== 多维评分过滤配置 =====================
# 是否启用多维评分过滤（语义+实体重叠+逻辑独特度+结构重要性），False时"按权重过滤"栏位填"权重过滤逻辑未启用"
SCORE_FILTER_ENABLED = True

# 条件-结论对评分权重（信息丰富的长文本，适合多维评分）
SCORE_WEIGHT_CC_SIM = 0.7          # 语义相关性（BGE-M3余弦相似度）
SCORE_WEIGHT_CC_OVERLAP = 0.1      # 实体重叠度（jieba分词+Jaccard相似系数）
SCORE_WEIGHT_CC_UNIQUENESS = 0.1   # 逻辑独特度（基于当前文章集的局部IDF）
SCORE_WEIGHT_CC_STRUCT = 0.1       # 结构重要性（检测数字、百分号、条款号等结构化特征）

# 政策场景评分权重（短文本2-6个汉字，不使用逻辑独特度）
SCORE_WEIGHT_SCENE_SIM = 0.8       # 语义相关性
SCORE_WEIGHT_SCENE_OVERLAP = 0.1   # 实体重叠度
SCORE_WEIGHT_SCENE_STRUCT = 0.1    # 结构重要性

# 时间约束评分权重（年份等结构信息很重要，不使用实体重叠）
SCORE_WEIGHT_TIME_SIM = 0.5        # 语义相关性
SCORE_WEIGHT_TIME_UNIQUENESS = 0.2 # 逻辑独特度
SCORE_WEIGHT_TIME_STRUCT = 0.3     # 结构重要性

# 概念断言关系类型优先级加分（断言评分 = 语义相似度 + 关系加分）
RELATION_BONUS = {
    "related_not_equal": 0.3,   # 相关但不可等同，信息量大
    "succession": 0.3,          # 替代关系，需要注意过渡期
    "mutually_exclusive": 0.2,  # 互斥关系，防止冲突
    "synonym": 0.0,             # 等价关系，信息量一般
    "property_of": 0.0          # 属性关系，信息量一般
}

# Top K 截断数量（每个类别按加权总分降序排序后最多保留的条数）
TOP_K_CC = 15        # 条件-结论对
TOP_K_SCENE = 10     # 政策场景
TOP_K_ASSERTION = 10  # 概念关系断言
TOP_K_TIME = 6       # 时间约束

# 法规文号前缀（用于结构重要性评分中匹配法规文号，如 国发[2007]39号、财税[2009]69号）
DOC_NUMBER_PREFIXES = ["财税", "国税", "税务总局", "公告", "国发"]

# 中文停用词表（用于jieba分词后过滤虚词和常见词）
STOPWORDS = set([
    "的", "了", "是", "在", "和", "与", "或", "以及", "按照", "根据",
    "对", "为", "由", "于", "之", "者", "被", "把", "将", "从", "到",
    "上", "下", "中", "有", "个", "这", "那", "不", "也", "都", "其",
    "等", "可", "应", "需", "要", "会", "能", "时", "如", "但", "并",
    "而", "及", "该", "此", "以", "当", "则", "若", "还", "已", "所",
    "优惠", "政策", "规定", "办法", "通知"
])


def _extract_article_summary(text):
    """从API返回内容中提取"文章汇总"章节，到---或下一个##标题为止"""
    import re
    # 匹配"文章汇总"章节：到---或下一个##标题或文本结束
    pattern = re.compile(r'##\s*文章汇总\s*\n(.*?)(?=\n---|\n##\s|\Z)', re.DOTALL)
    match = pattern.search(text)
    if match:
        return match.group(1).strip()
    return ""


def _fill_article_template(article_list_text):
    """将文章列表文本填充到模板中，替换{{}}占位符"""
    template_path = os.path.join(ROOT_DIR, 'docs', '提高回答质量', '提示词',
                                  '通过文章URL或Title获取文章内容_模板填充.md')
    with open(template_path, 'r', encoding='utf-8') as f:
        template = f.read()
    filled = template.replace('{{}}', article_list_text)
    return filled


def _get_articles_full(question: str, max_rounds: int = 12, system_prompt: str = "") -> list:
    """多轮对话获取文章，返回每轮内容的列表（不丢弃中间轮次的文章原文）
    question: 搜索关键词（B列提示词或A列问题）
    system_prompt: 可选，用户原始问题（A列），通过custom_system_prompt传给API
    """
    session_id = ""
    current_message = question
    all_parts = []

    for round_i in range(max_rounds):
        try:
            api_response, session_id = request_api(current_message, session_id, custom_system_prompt=system_prompt)
            content = parse_response(api_response)['full_content']
            if content and content.strip():
                all_parts.append(content)
                _log(f"[stability] 第{round_i + 1}轮获取完成（长度={len(content)}）")
            else:
                _log(f"[stability] 第{round_i + 1}轮返回为空")
            if is_confirmation_question(content):
                current_message = "同意，请继续，不需要调整。"
            elif is_incomplete_answer(content):
                current_message = "继续"
            else:
                break
        except Exception as e:
            _log(f"[stability] 第{round_i + 1}轮获取失败: {e}，继续下一轮")
            session_id = ""
            current_message = question

    _log(f"[stability] 文章获取完成，共{len(all_parts)}轮有内容，总长度={sum(len(p) for p in all_parts)}")
    return all_parts


# ===================== Combined Skill Prompt =====================

COMBINED_SKILL_SYSTEM = """你是一个税务多维度信息抽取专家。请从给定的用户问题和文章中，同时抽取以下四类信息，一次性输出。

输出格式：一个JSON对象，包含以下四个字段：

1. "condition_conclusion_pairs"：条件-结论对数组，每个对象包含：
   - condition: 触发条件（一句话，客观可判断）
   - conclusion: 对应的结论/规则
   - article_ids: 来源文章ID列表（如["ART_44382"]）
   要求：每条规则独立，不合并多个条件；例外或限制条件单独拆分；不要添加原文没有的信息。

2. "policy_scenes"：政策场景标签数组，每个元素是一个简短标签（如"西部大开发优惠"）。
   要求：标签来源于文章中的政策名称或常见税务术语；具有区分度，避免过于宽泛；不要遗漏。

3. "concept_relations"：概念关系断言数组，每个对象包含：
   - entity_a: 概念A名称
   - entity_b: 概念B名称
   - relation_type: 关系类型，严格限定为以下六种之一：
     hypernym(A是B的下位), synonym(完全等价), related_not_equal(相关但不可等同),
     mutually_exclusive(互斥不能同时适用), succession(B替代A可能带过渡期), property_of(A是B的属性)
   - evidence: 原文证据（摘录1-2句）
   - article_ids: 来源文章ID列表
   要求：只抽取明确体现的关系，每条必须有原文证据；特别关注跨文章之间的概念关系（替代、等价、互斥等）。

4. "time_constraints"：时间约束数组，每个对象包含：
   - policy_name: 政策名称或简称
   - constraint_type: "valid_for"(适用) / "invalid_for"(不适用) / "transitional"(过渡期保留)
   - condition: 具体条件描述
   - article_ids: 来源文章ID列表
   要求：只抽取明确的时间限定；政策废止但保留存量时拆分为invalid_for和transitional两条。

如果某类信息不存在，输出空数组[]。
特别要求：重点关注跨文章之间的关系——如果多篇文章涉及同一政策的不同阶段（如废止、替代、过渡期），务必在concept_relations和time_constraints中体现。
只输出JSON，不要其他解释。"""

FINAL_ANSWER_SYSTEM = "你是一个税务问答专家。请严格按以下结构和约束回答问题。"


# ===================== Helper Functions =====================

def _sse(data_dict):
    return f"data: {json.dumps(data_dict, ensure_ascii=False)}\n\n"


def _log(msg):
    """统一日志：同时输出到logger和VSCode控制台"""
    logger.info(msg)
    print(msg, flush=True)


def _call_llm(system_prompt, user_prompt, temperature=0.01, max_tokens=2000, timeout=300, max_retries=3):
    """Call LLM with system prompt and temperature control"""
    base_url = get_config('llm_base_url', '').rstrip('/')
    key = get_config('llm_api_key', '')
    model = get_config('model', '')

    if not base_url or not key or not model:
        raise ValueError("LLM API未配置（llm_base_url/llm_api_key/model）")

    url = base_url + '/v1/chat/completions'
    _log("_call_llm url:" + url)
    messages = []
    if system_prompt:
        messages.append({'role': 'system', 'content': system_prompt})
    messages.append({'role': 'user', 'content': user_prompt})

    body = {'model': model, 'messages': messages, 'temperature': temperature}
    if max_tokens > 0:
        body['max_tokens'] = max_tokens
    data = json.dumps(body, ensure_ascii=False).encode('utf-8')

    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {key}',
    }

    for attempt in range(1, max_retries + 1):
        try:
            resp = requests.post(url, data=data, headers=headers, timeout=timeout, verify=False)
            _log(f"[stability] _call_llm HTTP状态={resp.status_code} 响应长度={len(resp.text)}")
            resp.raise_for_status()
            try:
                r = resp.json()
            except Exception:
                _log(f"[stability] _call_llm JSON解析失败，原始响应前500字: {resp.text[:500]}")
                raise
            if 'error' in r:
                _log(f"[stability] _call_llm API返回错误: {r['error']}")
                if attempt >= max_retries:
                    raise RuntimeError(str(r['error']))
                time.sleep(min(attempt * 2, 10))
                continue
            content = r['choices'][0]['message']['content']
            _log(f"[stability] _call_llm 模型={model} temperature={temperature} 返回长度={len(content) if content else 0}")
            return content
        except Exception as e:
            _log(f"[stability] _call_llm 请求异常(第{attempt}次): {e}")
            if attempt >= max_retries:
                raise
            time.sleep(min(attempt * 2, 10))

    raise RuntimeError("LLM调用失败")


# ===================== BGE-M3 Embedding =====================

_bge_model = None

def _get_bge_model():
    """懒加载BGE-M3本地模型"""
    global _bge_model
    if _bge_model is None:
        from sentence_transformers import SentenceTransformer
        model_path = os.path.join(ROOT_DIR, "bge_m3_cache", "Xorbits", "bge-m3")
        _log(f"[stability] 正在加载BGE-M3模型（路径: {model_path}）...")
        _bge_model = SentenceTransformer(model_path)
        _log("[stability] BGE-M3模型加载完成")
    return _bge_model


def _generate_embedding(text):
    """用BGE-M3生成文本向量"""
    model = _get_bge_model()
    emb = model.encode(text[:500], normalize_embeddings=True)
    return emb.tolist()


# ===================== Article Grouping =====================

def _group_articles(article_parts, max_chars=None, sim_threshold=None):
    """按BGE-M3向量相似度对文章分组

    返回 list[list[str]]，每个内层列表是一组文章文本
    """
    import numpy as np

    max_chars = max_chars or get_config('group_max_chars', 12000)
    sim_threshold = sim_threshold or get_config('group_sim_threshold', 0.6)

    if len(article_parts) <= 1:
        return [article_parts] if article_parts else []

    # 为每篇文章生成向量
    article_data = []
    for i, content in enumerate(article_parts):
        vec = _generate_embedding(content)
        article_data.append({'content': content, 'vec': vec, 'len': len(content)})
    _log(f"[stability] 向量生成完成，共{len(article_data)}篇文章")

    groups = []  # list of {'articles': [...], 'vecs': [...], 'total_len': int}

    for art in article_data:
        # 单篇超长：独立成组
        if art['len'] > max_chars:
            groups.append({'articles': [art], 'vecs': [art['vec']], 'total_len': art['len']})
            continue

        placed = False
        for group in groups:
            # 字符数检查
            if group['total_len'] + art['len'] > max_chars:
                continue
            # 相似度检查：与组内已有文章的平均余弦相似度
            sims = [float(np.dot(art['vec'], v)) for v in group['vecs']]
            avg_sim = sum(sims) / len(sims)
            if avg_sim >= sim_threshold:
                group['articles'].append(art)
                group['vecs'].append(art['vec'])
                group['total_len'] += art['len']
                placed = True
                break

        if not placed:
            groups.append(
                {'articles': [art], 'vecs': [art['vec']], 'total_len': art['len']}
            )

    # 提取文章文本
    result = [[a['content'] for a in g['articles']] for g in groups]
    return result


# ===================== Merge Functions =====================

def _batch_embeddings(texts):
    """批量生成向量并缓存，避免重复调用BGE-M3"""
    cache = {}
    unique_texts = list(set(t for t in texts if t))
    if not unique_texts:
        return cache
    model = _get_bge_model()
    vecs = model.encode(unique_texts, normalize_embeddings=True, show_progress_bar=False)
    for text, vec in zip(unique_texts, vecs):
        cache[text] = vec.tolist()
    return cache


def _merge_condition_pairs(all_pairs, sim_threshold=None):
    """合并条件-结论对：语义相似度>=阈值则合并article_ids"""
    import numpy as np

    if not all_pairs:
        return []

    sim_threshold = sim_threshold or get_config('merge_cc_sim_threshold', 0.95)

    # 预计算所有文本的向量（只调一次BGE-M3）
    all_texts = [p.get('condition', '') + p.get('conclusion', '') for p in all_pairs]
    vec_cache = _batch_embeddings(all_texts)

    merged = []
    used_indices = set()

    for i, pair in enumerate(all_pairs):
        if i in used_indices:
            continue
        current = dict(pair)
        ids = list(current.get('article_ids', current.get('article_id', [])))
        if isinstance(ids, str):
            ids = [ids]

        text_i = all_texts[i]
        vec_i = vec_cache.get(text_i)
        if not vec_i:
            current['article_ids'] = ids
            if 'article_id' in current:
                del current['article_id']
            merged.append(current)
            continue

        for j in range(i + 1, len(all_pairs)):
            if j in used_indices:
                continue
            text_j = all_texts[j]
            vec_j = vec_cache.get(text_j)
            if vec_j:
                sim = float(np.dot(vec_i, vec_j))
                if sim >= sim_threshold:
                    other_ids = all_pairs[j].get('article_ids', all_pairs[j].get('article_id', []))
                    if isinstance(other_ids, str):
                        other_ids = [other_ids]
                    ids.extend([x for x in other_ids if x not in ids])
                    used_indices.add(j)

        current['article_ids'] = ids
        if 'article_id' in current:
            del current['article_id']
        merged.append(current)

    return merged


def _merge_time_constraints(all_constraints):
    """合并时间约束：同policy+同type+同条件则合并article_ids"""
    if not all_constraints:
        return []

    merged = []
    seen = {}
    for tc in all_constraints:
        key = (tc.get('policy_name', ''), tc.get('constraint_type', ''), tc.get('condition', ''))
        ids = list(tc.get('article_ids', tc.get('article_id', [])))
        if isinstance(ids, str):
            ids = [ids]
        if key in seen:
            idx = seen[key]
            existing_ids = merged[idx].get('article_ids', [])
            merged[idx]['article_ids'] = list(set(existing_ids + ids))
        else:
            seen[key] = len(merged)
            item = dict(tc)
            item['article_ids'] = ids
            if 'article_id' in item:
                del item['article_id']
            merged.append(item)

    return merged


def _parse_json_response(text):
    """Extract JSON from LLM response"""
    if not text:
        return None
    text = text.strip()
    if text.startswith('```'):
        text = text.split('\n', 1)[1] if '\n' in text else text[3:]
    if text.endswith('```'):
        text = text[:-3]
    text = text.strip()
    try:
        return json.loads(text)
    except Exception:
        pass
    for start_char, end_char in [('[', ']'), ('{', '}')]:
        start = text.find(start_char)
        end = text.rfind(end_char) + 1
        if start != -1 and end > start:
            try:
                return json.loads(text[start:end])
            except Exception:
                pass
    return None


# ===================== Post-Processing Filters =====================

def _compute_confidence(item):
    """计算置信度：基于source_count(多源验证)和evidence(证据质量)"""
    source_count = len(item.get('article_ids', []))
    conf = 0.9 if source_count >= 2 else 0.6
    evidence = item.get('evidence', '')
    if len(evidence) > 100 and ('第' in evidence or '条' in evidence):
        conf = min(1.0, conf + 0.1)
    return conf


def _filter_by_relevance(items, query_vec, text_fn, threshold):
    """相关性过滤：用BGE-M3计算余弦相似度，保留>=threshold的"""
    import numpy as np
    if not items:
        return []
    # 批量预计算所有文本向量
    texts = [text_fn(item) for item in items]
    vec_cache = _batch_embeddings(texts)
    kept = []
    for idx, item in enumerate(items):
        text = texts[idx]
        if not text:
            continue
        vec = vec_cache.get(text)
        if not vec:
            continue
        sim = float(np.dot(query_vec, vec))
        item['relevance_score'] = round(sim, 4)
        # 记录被过滤掉的item
        if sim < threshold:
            _log(f"[stability] 相关性过滤-丢弃: 相似度={item['relevance_score']} < {threshold}, 文本={text[:80]}")
        if sim >= threshold:
            kept.append(item)
    return kept


def _filter_by_confidence(items, min_conf):
    """置信度过滤：保留置信度>=min_conf的"""
    if not items:
        return []
    kept = []
    for item in items:
        conf = _compute_confidence(item)
        item['confidence'] = conf
        if conf >= min_conf:
            kept.append(item)
    return kept


def _is_specific(item):
    """判断断言是否具有专门性（针对特定主体）"""
    text = json.dumps(item, ensure_ascii=False)
    return any(kw in text for kw in SPECIFIC_KEYWORDS)


def _resolve_conflict_assertions(assertions):
    """概念关系冲突消解：按(entity_a, entity_b)分组，保留优先级最高的"""
    groups = {}
    for a in assertions:
        ea, eb = a.get('entity_a', ''), a.get('entity_b', '')
        key = tuple(sorted([ea, eb]))
        groups.setdefault(key, []).append(a)

    resolved = []
    for key, group in groups.items():
        if len(group) == 1:
            resolved.append(group[0])
        else:
            # 排序：发布时间晚 > 专门性高 > source_count大
            group.sort(key=lambda x: (
                x.get('publish_date', '1900-01-01'),
                _is_specific(x),
                len(x.get('article_ids', []))
            ), reverse=True)
            resolved.append(group[0])
            for discarded in group[1:]:
                _log(f"[stability] 冲突消解-丢弃断言: {discarded.get('entity_a', '')} "
                     f"↔ {discarded.get('entity_b', '')} ({discarded.get('relation_type', '')})")
    return resolved


def _resolve_conflict_cc_pairs(pairs):
    """条件-结论对冲突消解：同condition不同conclusion时保留优先级高的"""
    groups = {}
    for p in pairs:
        key = p.get('condition', '')
        groups.setdefault(key, []).append(p)

    resolved = []
    for key, group in groups.items():
        if len(group) == 1:
            resolved.append(group[0])
        else:
            group.sort(key=lambda x: (
                x.get('publish_date', '1900-01-01'),
                _is_specific(x),
                len(x.get('article_ids', []))
            ), reverse=True)
            resolved.append(group[0])
            for discarded in group[1:]:
                _log(f"[stability] 冲突消解-丢弃结论: condition={key[:50]}, "
                     f"conclusion={discarded.get('conclusion', '')[:50]}")
    return resolved


def _filter_skills_outputs(user_query, products):
    """三层过滤主函数：相关性 → 置信度 → 冲突消解"""
    import numpy as np

    _log(f"[stability] 开始三层过滤，用户问题: {user_query[:100]}")
    query_vec = _generate_embedding(user_query)

    # 1. 条件-结论对：相关性 → 置信度 → 冲突消解
    cc = list(products.get('condition_pairs', []))
    _log(f"[stability] 条件-结论对: 过滤前={len(cc)}")
    cc = _filter_by_relevance(cc, query_vec,
                              lambda p: p.get('condition', '') + ' → ' + p.get('conclusion', ''),
                              FILTER_REL_CC)
    _log(f"[stability] 条件-结论对: 相关性过滤后={len(cc)} (阈值={FILTER_REL_CC})")
    cc = _filter_by_confidence(cc, FILTER_CONF_CC)
    _log(f"[stability] 条件-结论对: 置信度过滤后={len(cc)} (阈值={FILTER_CONF_CC})")
    cc = _resolve_conflict_cc_pairs(cc)
    _log(f"[stability] 条件-结论对: 冲突消解后={len(cc)}")

    # 2. 政策场景：仅相关性过滤
    scenes = list(products.get('scene_enum', []))
    scene_dicts = [{'label': s} for s in scenes]
    _log(f"[stability] 政策场景: 过滤前={len(scene_dicts)}")
    scene_dicts = _filter_by_relevance(scene_dicts, query_vec,
                                       lambda s: s.get('label', ''), FILTER_REL_SCENE)
    filtered_scenes = [s['label'] for s in scene_dicts]
    _log(f"[stability] 政策场景: 相关性过滤后={len(filtered_scenes)} (阈值={FILTER_REL_SCENE})")

    # 3. 概念关系：相关性 → 置信度 → 冲突消解
    rels = list(products.get('assertions_raw', []))
    _log(f"[stability] 概念断言: 过滤前={len(rels)}")
    rels = _filter_by_relevance(rels, query_vec,
                                lambda r: f"{r.get('entity_a', '')} {r.get('entity_b', '')} {r.get('relation_type', '')}",
                                FILTER_REL_ASSERTION)
    _log(f"[stability] 概念断言: 相关性过滤后={len(rels)} (阈值={FILTER_REL_ASSERTION})")
    rels = _filter_by_confidence(rels, FILTER_CONF_ASSERTION)
    _log(f"[stability] 概念断言: 置信度过滤后={len(rels)} (阈值={FILTER_CONF_ASSERTION})")
    rels = _resolve_conflict_assertions(rels)
    _log(f"[stability] 概念断言: 冲突消解后={len(rels)}")

    # 4. 时间约束：相关性 → 置信度
    tcs = list(products.get('time_constraints', []))
    _log(f"[stability] 时间约束: 过滤前={len(tcs)}")
    tcs = _filter_by_relevance(tcs, query_vec,
                               lambda t: f"{t.get('policy_name', '')} {t.get('constraint_type', '')} {t.get('condition', '')}",
                               FILTER_REL_TIME)
    _log(f"[stability] 时间约束: 相关性过滤后={len(tcs)} (阈值={FILTER_REL_TIME})")
    tcs = _filter_by_confidence(tcs, FILTER_CONF_TIME)
    _log(f"[stability] 时间约束: 置信度过滤后={len(tcs)} (阈值={FILTER_CONF_TIME})")

    return {
        'condition_pairs': cc,
        'policy_scenes': filtered_scenes,
        'concept_relations': rels,
        'time_constraints': tcs
    }



    if not text:
        return None
    text = text.strip()
    if text.startswith('```'):
        text = text.split('\n', 1)[1] if '\n' in text else text[3:]
    if text.endswith('```'):
        text = text[:-3]
    text = text.strip()
    try:
        return json.loads(text)
    except Exception:
        pass
    for start_char, end_char in [('[', ']'), ('{', '}')]:
        start = text.find(start_char)
        end = text.rfind(end_char) + 1
        if start != -1 and end > start:
            try:
                return json.loads(text[start:end])
            except Exception:
                pass
    return None


def _parse_articles_from_text(text):
    """将文章内容按标记拆分为逐篇文章列表

    支持两种格式：
      ## {序号}. NTPSID: {数字id}
      ## 文章{序号}：{标题}
    以 --- 或不同的文章标题作为分隔
    """
    import re
    if not text or not text.strip():
        return []

    # 按 --- 分割（精确匹配单独一行的三个连字符）
    separator_pattern = re.compile(r'^---$', re.MULTILINE)
    segments = separator_pattern.split(text)

    # 两种文章标题格式
    ntpsid_pattern = re.compile(r'^##\s*\d+\.\s*NTPSID:\s*\d+', re.MULTILINE)
    article_title_pattern = re.compile(r'^##\s*文章\d+[：:]', re.MULTILINE)

    articles = []

    for segment in segments:
        segment = segment.strip()
        if not segment:
            continue

        # 优先按NTPSID标记拆分
        matches = list(ntpsid_pattern.finditer(segment))
        if not matches:
            # 其次按"文章N："标记拆分
            matches = list(article_title_pattern.finditer(segment))

        if not matches:
            # 没有任何标记，整段作为一篇文章
            articles.append(segment)
            continue

        for i, m in enumerate(matches):
            start = m.start()
            end = matches[i + 1].start() if i + 1 < len(matches) else len(segment)
            article_text = segment[start:end].strip()
            if article_text:
                articles.append(article_text)

    # 如果没拆出任何文章，返回原文作为单篇
    if not articles:
        articles = [text.strip()]

    return articles


def _dedupe_strings(items):
    """去重字符串列表，保持顺序"""
    seen = set()
    result = []
    for item in items:
        s = str(item) if not isinstance(item, str) else item
        if s not in seen:
            seen.add(s)
            result.append(s)
    return result


# ===================== Assertion Validation (C1-C7) =====================

def _validate_assertions(assertions):
    """Validate and clean concept relation assertions (rules C1-C7)"""
    if not assertions or not isinstance(assertions, list):
        return []

    cleaned = [dict(a) for a in assertions if isinstance(a, dict)]

    # C1: 删除自引用（entity_a 与 entity_b 相同时为无意义断言）
    cleaned = [a for a in cleaned if a.get('entity_a', '') != a.get('entity_b', '')]

    # C2: 去重合并（相同三元组合并为一条，拼接 evidence）
    seen = {}
    for a in cleaned:
        key = (a.get('entity_a', ''), a.get('entity_b', ''), a.get('relation_type', ''))
        if key in seen:
            old_ev = seen[key].get('evidence', '')
            new_ev = a.get('evidence', '')
            if new_ev and new_ev not in old_ev:
                seen[key]['evidence'] = (old_ev + '；' + new_ev) if old_ev else new_ev
        else:
            seen[key] = dict(a)
    cleaned = list(seen.values())

    # C3: 同义+互斥冲突（同一对概念既是 synonym 又是 mutually_exclusive 时，删除互斥）
    pair_rels = {}
    for a in cleaned:
        k = (a.get('entity_a', ''), a.get('entity_b', ''))
        pair_rels.setdefault(k, set()).add(a.get('relation_type', ''))

    conflict_c3 = {k for k, rels in pair_rels.items() if 'synonym' in rels and 'mutually_exclusive' in rels}
    if conflict_c3:
        cleaned = [a for a in cleaned if not (
            (a.get('entity_a', ''), a.get('entity_b', '')) in conflict_c3
            and a.get('relation_type', '') == 'mutually_exclusive'
        )]

    # C4: 同义+下位冲突（同一对概念既是 synonym 又是 hypernym 时，删除下位）
    conflict_c4 = {k for k, rels in pair_rels.items() if 'synonym' in rels and 'hypernym' in rels}
    if conflict_c4:
        cleaned = [a for a in cleaned if not (
            (a.get('entity_a', ''), a.get('entity_b', '')) in conflict_c4
            and a.get('relation_type', '') == 'hypernym'
        )]

    # C5: 互斥对称补全（A 互斥 B 时自动补充 B 互斥 A）
    me_pairs = {(a.get('entity_a', ''), a.get('entity_b', ''))
                for a in cleaned if a.get('relation_type', '') == 'mutually_exclusive'}
    existing = {(a.get('entity_a', ''), a.get('entity_b', '')) for a in cleaned}
    for ea, eb in me_pairs:
        if (eb, ea) not in existing:
            cleaned.append({
                'entity_a': eb, 'entity_b': ea,
                'relation_type': 'mutually_exclusive',
                'evidence': 'derived: symmetric complement',
                'derived': True
            })

    # C6: 证据实体匹配（evidence 中必须包含 entity_a 或 entity_b 的原文，否则删除）— 已禁用
    # validated = []
    # for a in cleaned:
    #     if a.get('derived'):
    #         validated.append(a)
    #         continue
    #     ev = a.get('evidence', '')
    #     ea = a.get('entity_a', '')
    #     eb = a.get('entity_b', '')
    #     if (ea and ea in ev) or (eb and eb in ev):
    #         validated.append(a)
    # cleaned = validated

    # C7: 证据长度截断（evidence 超过 500 字符时截断，保留完整内容用于Excel输出，截断版本仅用于内部比对）
    # 已禁用截断，Excel中完整显示
    # for a in cleaned:
    #     if len(a.get('evidence', '')) > 500:
    #         a['evidence'] = a['evidence'][:200] + '...'

    return cleaned


def _convert_constraints_to_text(cleaned_assertions):
    """Convert cleaned assertions to natural language constraint texts"""
    texts = []
    LQ = '“'  # left double quotation mark
    RQ = '”'  # right double quotation mark
    for a in cleaned_assertions:
        rt = a.get('relation_type', '')
        ea = a.get('entity_a', '')
        eb = a.get('entity_b', '')
        if rt == 'hypernym':
            texts.append(f'{LQ}{ea}{RQ}是{LQ}{eb}{RQ}的下位概念。禁止将二者视为并列或等同。')
        elif rt == 'synonym':
            texts.append(f'{LQ}{ea}{RQ}与{LQ}{eb}{RQ}为同义概念，回答时可直接等同替换。')
        elif rt == 'related_not_equal':
            texts.append(f'{LQ}{ea}{RQ}与{LQ}{eb}{RQ}相关但不可等同。禁止在推理中将二者相互替换或视为相同。')
        elif rt == 'mutually_exclusive':
            texts.append(f'{LQ}{ea}{RQ}与{LQ}{eb}{RQ}互斥，不能同时适用。')
        elif rt == 'succession':
            texts.append(f'{LQ}{eb}{RQ}替代{LQ}{ea}{RQ}，但注意是否存在过渡期保留（见时间约束）。')
        elif rt == 'property_of':
            texts.append(f'{LQ}{ea}{RQ}是{LQ}{eb}{RQ}的一个属性/参数，不等同于整体政策。')
    return texts


# ===================== 多维评分过滤函数 =====================

def _tokenize(text):
    """jieba分词+停用词过滤，返回set（用于Jaccard计算）"""
    import jieba
    words = jieba.lcut(text)
    return set(w for w in words if w not in STOPWORDS and (len(w) > 1 or w.isdigit()))


def _tokenize_list(text):
    """jieba分词+停用词过滤，返回list（保留词频，用于IDF加权平均计算）"""
    import jieba
    words = jieba.lcut(text)
    return [w for w in words if w not in STOPWORDS and (len(w) > 1 or w.isdigit())]


def _jaccard_similarity(set1, set2):
    """计算两个集合的Jaccard相似系数"""
    if not set1 and not set2:
        return 0.0
    inter = len(set1 & set2)
    union = len(set1 | set2)
    return inter / union if union > 0 else 0.0


def _build_local_idf(article_texts):
    """基于当前文章集构建局部IDF字典
    article_texts: list[str]，每篇文章的文本
    返回: dict{词: idf值}
    """
    import math
    from collections import Counter
    N = len(article_texts)
    if N == 0:
        return {}
    df = Counter()
    for text in article_texts:
        words = _tokenize(text)
        for w in words:
            df[w] += 1
    idf = {}
    for w, doc_cnt in df.items():
        val = math.log(N / (1 + doc_cnt))
        idf[w] = max(val, 0.0)
    return idf


def _compute_avg_idf(product_text, idf_dict):
    """计算产物文本的平均IDF（逻辑独特度）"""
    words_list = _tokenize_list(product_text)
    if not words_list:
        return 0.0
    total = sum(idf_dict.get(w, 0.0) for w in words_list)
    return total / len(words_list)


def _structural_score(text):
    """结构重要性评分：检测百分号、条款号、年份、法规文号、金额等"""
    import re
    score = 0.0
    if re.search(r'\d+(?:\.\d+)?%', text):
        score += 0.3
    if re.search(r'第[零一二三四五六七八九十百千万\d]+条', text):
        score += 0.3
    if re.search(r'\d{4}年', text):
        score += 0.2
    if re.search(r'(' + '|'.join(DOC_NUMBER_PREFIXES) + r')[\[\(]\d{4}[\d\]\)]', text):
        score += 0.2
    if re.search(r'[\[\(]\d{4}[\]\)]\s*\d+号', text):
        score += 0.2
    if re.search(r'\d{1,3}(?:万|亿)?元', text):
        score += 0.1
    return min(score, 1.0)


def _normalize_scores(scores):
    """线性归一化到[0,1]，所有值相同时返回0.5"""
    if not scores:
        return []
    min_s = min(scores)
    max_s = max(scores)
    if max_s == min_s:
        return [0.5] * len(scores)
    return [(s - min_s) / (max_s - min_s) for s in scores]


def _score_filter_products(user_query, article_texts, products):
    """多维评分过滤主函数：分类型评分→排序截断TopK
    每种产物类型使用不同的评分公式和权重：
    - 条件-结论对：语义+重叠+独特度+结构
    - 概念断言：语义相似度+关系类型加分
    - 时间约束：语义+独特度+结构（无重叠）
    - 政策场景：语义+重叠+结构（无独特度）
    返回 dict: {condition_pairs, policy_scenes, concept_relations, time_constraints}
    """
    import numpy as np

    _log(f"[stability] 开始多维评分过滤，用户问题: {user_query[:100]}")

    # 构建局部IDF
    idf_dict = _build_local_idf(article_texts)
    _log(f"[stability] 局部IDF构建完成，词典大小={len(idf_dict)}")

    # 用户问题的分词和向量（只算一次）
    q_tokens = _tokenize(user_query)
    query_vec = _generate_embedding(user_query)

    # ---- 通用多维评分函数 ----
    def _score_multi_dim(items, text_fn, weights, top_k, label):
        """多维评分：仅计算权重非0的维度，仅对IDF归一化→加权总分→排序截断"""
        if not items:
            _log(f"[stability] {label}: 无数据，跳过")
            return []

        _log(f"[stability] {label}: 评分前={len(items)}条")

        w_sim = weights.get('sim', 0)
        w_overlap = weights.get('overlap', 0)
        w_uniqueness = weights.get('uniqueness', 0)
        w_struct = weights.get('struct', 0)

        # 批量预计算向量
        texts = [text_fn(item) for item in items]
        vec_cache = _batch_embeddings(texts)

        sim_scores, overlap_scores, uniqueness_scores, struct_scores = [], [], [], []

        for i, (item, text) in enumerate(zip(items, texts)):
            if w_sim > 0:
                vec = vec_cache.get(text)
                sim = float(np.dot(query_vec, vec)) if vec else 0.0
                sim_scores.append(sim)
            if w_overlap > 0:
                p_tokens = _tokenize(text)
                overlap = _jaccard_similarity(q_tokens, p_tokens)
                overlap_scores.append(overlap)
            if w_uniqueness > 0:
                avg_idf = _compute_avg_idf(text, idf_dict)
                uniqueness_scores.append(avg_idf)
            if w_struct > 0:
                struct = _structural_score(text)
                struct_scores.append(struct)

        # 仅IDF需要归一化（其他维度已在[0,1]范围内）
        uniqueness_norm = _normalize_scores(uniqueness_scores) if w_uniqueness > 0 else []

        for i, item in enumerate(items):
            total = 0.0
            if w_sim > 0:
                item['score_sim'] = round(sim_scores[i], 4)
                total += w_sim * sim_scores[i]
            if w_overlap > 0:
                item['score_overlap'] = round(overlap_scores[i], 4)
                total += w_overlap * overlap_scores[i]
            if w_uniqueness > 0:
                item['score_uniqueness'] = round(uniqueness_scores[i], 4)
                total += w_uniqueness * uniqueness_norm[i]
            if w_struct > 0:
                item['score_struct'] = round(struct_scores[i], 4)
                total += w_struct * struct_scores[i]
            item['total_score'] = round(total, 4)

        sorted_items = sorted(items, key=lambda x: x['total_score'], reverse=True)
        kept = sorted_items[:top_k]

        _log(f"[stability] {label}: 评分后保留={len(kept)}条 (TopK={top_k})")
        for rank, k in enumerate(sorted_items, 1):
            status = "保留" if rank <= top_k else "截断"
            _log(f"[stability]   [{rank}] {status} 总分={k['total_score']} "
                 f"(语义={k.get('score_sim', '-')}, 重叠={k.get('score_overlap', '-')}, "
                 f"独特={k.get('score_uniqueness', '-')}, 结构={k.get('score_struct', '-')}) "
                 f"文本={text_fn(k)}")

        return kept

    result = {}

    # ---- 1. 条件-结论对（多维评分：语义+重叠+独特度+结构） ----
    cc_items = [dict(p) for p in products.get('condition_pairs', [])]
    result['condition_pairs'] = _score_multi_dim(
        cc_items,
        lambda p: p.get('condition', '') + ' → ' + p.get('conclusion', ''),
        {'sim': SCORE_WEIGHT_CC_SIM, 'overlap': SCORE_WEIGHT_CC_OVERLAP,
         'uniqueness': SCORE_WEIGHT_CC_UNIQUENESS, 'struct': SCORE_WEIGHT_CC_STRUCT},
        TOP_K_CC, '条件-结论对')

    # ---- 2. 概念关系断言（语义相似度 + 关系类型加分） ----
    rel_items = [dict(r) for r in products.get('assertions_raw', [])]
    if not rel_items:
        _log(f"[stability] 概念断言: 无数据，跳过")
        result['concept_relations'] = []
    else:
        _log(f"[stability] 概念断言: 评分前={len(rel_items)}条")
        _assertion_text_fn = lambda r: (f"{r.get('entity_a', '')} {r.get('entity_b', '')}"
                                         + (f" {r.get('evidence', '')}" if r.get('evidence') else ''))
        rel_texts = [_assertion_text_fn(r) for r in rel_items]
        rel_vec_cache = _batch_embeddings(rel_texts)

        for item, text in zip(rel_items, rel_texts):
            vec = rel_vec_cache.get(text)
            sem = float(np.dot(query_vec, vec)) if vec else 0.0
            bonus = RELATION_BONUS.get(item.get('relation_type', ''), 0.0)
            item['score_sim'] = round(sem, 4)
            item['score_bonus'] = round(bonus, 4)
            item['total_score'] = round(sem + bonus, 4)

        sorted_rels = sorted(rel_items, key=lambda x: x['total_score'], reverse=True)
        rel_kept = sorted_rels[:TOP_K_ASSERTION]

        _log(f"[stability] 概念断言: 评分后保留={len(rel_kept)}条 (TopK={TOP_K_ASSERTION})")
        for rank, k in enumerate(sorted_rels, 1):
            status = "保留" if rank <= TOP_K_ASSERTION else "截断"
            _log(f"[stability]   [{rank}] {status} 总分={k['total_score']} "
                 f"(语义={k['score_sim']}, 加分={k['score_bonus']}, "
                 f"类型={k.get('relation_type', '')}) 文本={_assertion_text_fn(k)}")

        result['concept_relations'] = rel_kept

    # ---- 3. 时间约束（多维评分：语义+独特度+结构，不使用重叠） ----
    tc_items = [dict(t) for t in products.get('time_constraints', [])]
    result['time_constraints'] = _score_multi_dim(
        tc_items,
        lambda t: f"{t.get('policy_name', '')} {t.get('constraint_type', '')} {t.get('condition', '')}",
        {'sim': SCORE_WEIGHT_TIME_SIM, 'uniqueness': SCORE_WEIGHT_TIME_UNIQUENESS,
         'struct': SCORE_WEIGHT_TIME_STRUCT},
        TOP_K_TIME, '时间约束')

    # ---- 4. 政策场景（多维评分：语义+重叠+结构，不使用独特度） ----
    scenes_raw = products.get('scene_enum', [])
    if not scenes_raw:
        _log(f"[stability] 政策场景: 无数据，跳过")
        result['policy_scenes'] = []
    else:
        scene_items = [{'label': s, '_original': s} for s in scenes_raw]
        scene_kept = _score_multi_dim(
            scene_items,
            lambda s: s['_original'],
            {'sim': SCORE_WEIGHT_SCENE_SIM, 'overlap': SCORE_WEIGHT_SCENE_OVERLAP,
             'struct': SCORE_WEIGHT_SCENE_STRUCT},
            TOP_K_SCENE, '政策场景')
        result['policy_scenes'] = [s['_original'] for s in scene_kept]

    return result


def _assemble_final_prompt(user_query, condition_pairs, scene_enum, constraint_texts, time_constraints):
    """以Prompt.md为模板，将4个Skill产物填入对应章节后返回完整提示词"""

    # Scene enumeration
    if isinstance(scene_enum, list) and scene_enum:
        scene_list = "\n".join(f"- {s}" if isinstance(s, str) else f"- {s}" for s in scene_enum)
    else:
        scene_list = "未提取到政策场景"

    # Constraint texts
    constraints = "\n".join(f"- {t}" for t in constraint_texts) if constraint_texts else "无特殊概念约束"

    # Condition-conclusion pairs
    if isinstance(condition_pairs, list) and condition_pairs:
        cc_lines = []
        for cc in condition_pairs:
            cond = cc.get('condition', '')
            conc = cc.get('conclusion', '')
            cc_lines.append(f"- 条件：{cond} → 结论：{conc}")
        cc_text = "\n".join(cc_lines)
    else:
        cc_text = "未抽取到条件-结论对"

    # Time constraints
    if isinstance(time_constraints, list) and time_constraints:
        tc_lines = []
        for tc in time_constraints:
            ct = tc.get('constraint_type', '')
            cond = tc.get('condition', '')
            name = tc.get('policy_name', '')
            labels = {'valid_for': '适用', 'invalid_for': '不适用', 'transitional': '过渡期'}
            tc_lines.append(f"- {name}：{labels.get(ct, ct)} — {cond}")
        time_text = "\n".join(tc_lines)
    else:
        time_text = "无特殊时间约束"

    # 读取Prompt.md模板，按{{}}分割后填入4个Skill产物
    template_path = os.path.join(ROOT_DIR, 'Prompt.md')
    with open(template_path, 'r', encoding='utf-8') as f:
        template = f.read()

    parts = template.split('{{}}')
    # parts: [模板前半(到section5之前), section5-6之间, section6-7之间, section7-8之间, 模板后半(section8之后)]
    skill_outputs = [scene_list, constraints, cc_text, time_text]

    filled = parts[0]
    for i, output in enumerate(skill_outputs):
        filled += output
        filled += parts[i + 1]

    return filled


# ===================== Routes =====================

@stability_bp.route('/stability_upload', methods=['POST'])
def stability_upload():
    """上传Excel，解析A列(问题)、B列(提示词)、C列(文章内容)"""
    file = flask_request.files.get('file')
    if not file or not allowed_file(file.filename):
        return jsonify({'error': '请上传 .xlsx 或 .xls 文件'}), 400

    import re
    safe_name = re.sub(r'[\\/*?:"<>|]', '_', file.filename)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{timestamp}_{safe_name}"
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        questions = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            question = str(row[0]).strip() if len(row) > 0 and row[0] else ''
            prompt = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            articles_text = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            if not question or question == 'None':
                continue
            has_articles = bool(articles_text)
            article_count = 0
            if has_articles:
                parsed = _parse_articles_from_text(articles_text)
                article_count = len(parsed)
            questions.append({
                'row': row_idx,
                'question': question,
                'prompt': prompt,
                'has_prompt': bool(prompt),
                'has_articles': has_articles,
                'article_count': article_count,
                'articles_preview': articles_text[:200] + '...' if len(articles_text) > 200 else articles_text
            })
        wb.close()
    except Exception as e:
        return jsonify({'error': f'解析Excel失败: {e}'}), 400

    if not questions:
        return jsonify({'error': '未找到有效数据（A列需有问题）'}), 400

    _log(f"[stability] Excel解析完成: {filename}, 共{len(questions)}题")
    prompt_count = 0
    articles_count = 0
    for q in questions:
        _log(f"[stability]   行{q['row']}: 问题={q['question'][:80]}{'...' if len(q['question']) > 80 else ''}")
        if q['has_prompt']:
            prompt_count += 1
        if q['has_articles']:
            articles_count += 1
            _log(f"[stability]     C列文章: {q['article_count']}篇")
    _log(f"[stability] 共读取{len(questions)}个问题，其中{prompt_count}个有提示词，{articles_count}个有C列文章")
    
    return jsonify({
        'filename': filename,
        'question_count': len(questions),
        'questions': questions
    })


@stability_bp.route('/stability_process', methods=['POST'])
def stability_process():
    """SSE流式处理：获取文章→四Skill并行→校验→组装→生成最终回答"""
    data = flask_request.get_json()
    filename = data.get('filename', '')
    thread_count = max(1, int(data.get('thread_count', 2)))

    filepath = os.path.join(UPLOAD_FOLDER, filename)
    if not os.path.exists(filepath):
        return jsonify({'error': '文件不存在，请重新上传'}), 400

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        questions = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            question = str(row[0]).strip() if len(row) > 0 and row[0] else ''
            prompt = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            articles_text = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            if not question or question == 'None':
                continue
            questions.append((row_idx, question, prompt, articles_text))
        wb.close()
    except Exception as e:
        return jsonify({'error': f'解析Excel失败: {e}'}), 400

    total = len(questions)
    if total == 0:
        return jsonify({'error': '未找到有效问题'}), 400

    model_name = get_config('model', '')
    _log(f"[stability] 开始处理 {total}题 | {thread_count}线程 | 模型={model_name}")

    def generate():
        all_results = []

        for q_idx, (row_num, question, kb_prompt, c_articles) in enumerate(questions):
            _log(f"[stability] ===== 第{q_idx + 1}/{total}题 (行{row_num}) =====")
            yield _sse({'type': 'question_start', 'question_idx': q_idx, 'total_questions': total,
                        'question': question, 'prompt': kb_prompt})

            products = {
                'condition_pairs': [],
                'scene_enum': [],
                'assertions_raw': [],
                'assertions_cleaned': [],
                'constraint_texts': [],
                'time_constraints': [],
                'articles_text': '',
                'final_prompt': '',
                'final_answer': ''
            }

            # ---- Step 1: Get articles ----
            if c_articles:
                # C列有文章内容，直接拆分使用，跳过知识库API
                _log(f"\n{'='*60}\n[stability] ----------------Step 1: 从Excel C列读取文章---------------------\n{'='*60}")
                _log(f"[stability] row={row_num} 从Excel C列读取文章（长度={len(c_articles)}）")
                yield _sse({'type': 'step_start', 'question_idx': q_idx,
                            'step_id': 'articles', 'step_label': '读取文章内容(Excel C列)',
                            'system_prompt': '(从Excel C列直接读取)', 'user_prompt': c_articles[:500]})
                article_parts = _parse_articles_from_text(c_articles)
                articles_full_text = '\n\n'.join(article_parts)
                products['articles_text'] = articles_full_text
                for pi, part in enumerate(article_parts):
                    _log(f"[stability] row={row_num} 文章{pi+1}（长度={len(part)}）: {part[:200]}{'...' if len(part) > 200 else ''}")
                _log(f"[stability] row={row_num} 共{len(article_parts)}篇文章，总长度={len(articles_full_text)}")
                yield _sse({'type': 'step_complete', 'question_idx': q_idx,
                            'step_id': 'articles', 'step_label': '读取文章内容(Excel C列)',
                            'response': articles_full_text})
            else:
                # C列为空，通过两步API获取文章内容
                yield _sse({'type': 'step_start', 'question_idx': q_idx,
                            'step_id': 'articles', 'step_label': '获取文章内容(知识库API)',
                            'system_prompt': '(通过5007知识库API获取)', 'user_prompt': kb_prompt or question})
                try:
                    # ---- Step 1a: 第一次API调用，获取文章列表 ----
                    _log(f"\n{'='*60}\n[stability] ----------------Step 1a: 第一次API调用，获取文章列表---------------------\n{'='*60}")
                    _log(f"[stability] row={row_num} message={question}, system_prompt={kb_prompt or question}")
                    raw_parts = _get_articles_full(question, system_prompt=kb_prompt or question)
                    raw_text = '\n\n'.join(raw_parts)
                    _log(f"[stability] row={row_num} Step1a API返回内容（长度={len(raw_text)}）:\n{raw_text}")

                    # ---- Step 1b: 提取"文章汇总"章节 ----
                    _log(f"\n{'='*60}\n[stability] ----------------Step 1b: 提取文章汇总章节---------------------\n{'='*60}")
                    article_summary = _extract_article_summary(raw_text)
                    _log(f"[stability] row={row_num} Step1b 提取到文章汇总（长度={len(article_summary)}）:\n{article_summary}")

                    if not article_summary:
                        _log(f"[stability] row={row_num} Step1b 未找到文章汇总章节，使用原始返回内容")
                        article_parts = []
                        for part in raw_parts:
                            article_parts.extend(_parse_articles_from_text(part))
                    else:
                        # ---- Step 1c: 填充模板，第二次API调用获取文章原文 ----
                        _log(f"\n{'='*60}\n[stability] ----------------Step 1c: 填充模板 + 第二次API调用获取文章原文---------------------\n{'='*60}")
                        filled_prompt = _fill_article_template(article_summary)
                        _log(f"[stability] row={row_num} Step1c 填充后的模板提示词（长度={len(filled_prompt)}）:\n{filled_prompt}")

                        yield _sse({'type': 'step_start', 'question_idx': q_idx,
                                    'step_id': 'articles_fetch', 'step_label': '获取文章原文(第二次API)',
                                    'system_prompt': filled_prompt[:300], 'user_prompt': question[:200]})

                        _log(f"[stability] row={row_num} Step1c 第二次API调用，获取文章原文")
                        fetch_parts = _get_articles_full("获取文章内容", system_prompt=filled_prompt)
                        fetch_text = '\n\n'.join(fetch_parts)
                        _log(f"[stability] row={row_num} Step1c 第二次API返回内容（长度={len(fetch_text)}）:\n{fetch_text}")

                        # 拆分文章
                        _log(f"\n{'='*60}\n[stability] ----------------Step 1d: 文章拆分---------------------\n{'='*60}")
                        article_parts = []
                        for part in fetch_parts:
                            article_parts.extend(_parse_articles_from_text(part))
                        _log(f"[stability] row={row_num} Step1d 拆分得到{len(article_parts)}篇文章")

                    _log(f"\n{'='*60}\n[stability] ----------------Step 1 完成: 文章汇总---------------------\n{'='*60}")
                    articles_full_text = '\n\n'.join(article_parts)
                    products['articles_text'] = articles_full_text
                    for pi, part in enumerate(article_parts):
                        _log(f"[stability] row={row_num} 文章{pi+1}（长度={len(part)}）: {part[:200]}{'...' if len(part) > 200 else ''}")
                    _log(f"[stability] row={row_num} 共获取{len(article_parts)}篇文章，总长度={len(articles_full_text)}")
                    yield _sse({'type': 'step_complete', 'question_idx': q_idx,
                                'step_id': 'articles', 'step_label': '获取文章内容(知识库API)',
                                'response': articles_full_text})
                except Exception as e:
                    _log(f"[stability] row={row_num} 文章获取失败: {e}")
                    yield _sse({'type': 'step_error', 'question_idx': q_idx,
                                'step_id': 'articles', 'step_label': '获取文章内容(知识库API)',
                                'error': str(e)})
                    article_parts = [kb_prompt or question]
                    articles_full_text = article_parts[0]
                    products['articles_text'] = articles_full_text

            # ---- Step 2a: Group articles ----
            _log(f"\n{'='*60}\n[stability] ----------------Step 2a: BGE-M3向量文章分组---------------------\n{'='*60}")
            max_chars = get_config('group_max_chars', 12000)
            sim_threshold = get_config('group_sim_threshold', 0.6)
            _log(f"[stability] row={row_num} 开始分组（max_chars={max_chars}, sim_threshold={sim_threshold}）")
            yield _sse({'type': 'step_start', 'question_idx': q_idx,
                        'step_id': 'grouping', 'step_label': '文章分组(向量相似度)',
                        'system_prompt': f'(BGE-M3向量+余弦相似度>={sim_threshold})',
                        'user_prompt': f'{len(article_parts)}篇文章'})

            try:
                groups = _group_articles(article_parts, max_chars, sim_threshold)
            except Exception as e:
                _log(f"[stability] row={row_num} 分组失败，回退为逐篇处理: {e}")
                groups = [[art] for art in article_parts]

            group_info = ' | '.join(f"组{i+1}:{len(g)}篇" for i, g in enumerate(groups))
            _log(f"[stability] row={row_num} 分组完成: {len(groups)}组 — {group_info}")
            for gi, g in enumerate(groups):
                import re
                art_ids = []
                for art in g:
                    m = re.search(r'NTPS\s*ID[：:]*\s*(\d+)', art, re.IGNORECASE)
                    art_ids.append(f"ART_{m.group(1)}" if m else f"(未知)")
                total_chars = sum(len(a) for a in g)
                _log(f"[stability]   组{gi+1}: 文章={art_ids}, 总字符数={total_chars}")
            yield _sse({'type': 'step_complete', 'question_idx': q_idx,
                        'step_id': 'grouping', 'step_label': '文章分组(向量相似度)',
                        'response': f"分为{len(groups)}组: {group_info}"})

            # 2b: 每组调用合并Skill（并行）
            _log(f"\n{'='*60}\n[stability] ----------------Step 2b: 并行LLM抽取四类信息---------------------\n{'='*60}")
            task_queue = queue.Queue()

            def _run_combined(group_idx, group_articles):
                label = f"合并Skill(组{group_idx+1})"
                sid = f'combined_group{group_idx+1}'
                try:
                    # 为每篇文章提取NTPSID作为ID
                    import re
                    arts_text_parts = []
                    group_art_ids = []
                    for ai, art in enumerate(group_articles):
                        ntpsid_match = re.search(r'NTPS\s*ID[：:]*\s*(\d+)', art, re.IGNORECASE)
                        art_id = f"ART_{ntpsid_match.group(1)}" if ntpsid_match else f"ART_{row_num}_G{group_idx+1}_P{ai+1}"
                        group_art_ids.append(art_id)
                        arts_text_parts.append(f"### 文章ID: {art_id}\n{art}")
                    combined_text = '\n\n'.join(arts_text_parts)
                    user_prompt = f"用户问题：{question}\n\n文章列表：\n{combined_text}"
                    resp = _call_llm(COMBINED_SKILL_SYSTEM, user_prompt,
                                     temperature=0, max_tokens=0, timeout=600)
                    parsed = _parse_json_response(resp)
                    if parsed is None and resp:
                        _log(f"[stability] row={row_num} {label} JSON解析失败，原始响应前500字: {resp[:500]}")
                    task_queue.put((sid, label, resp, parsed, None, group_art_ids))
                except Exception as e:
                    task_queue.put((sid, label, None, None, str(e), []))

            # Send start events + submit tasks
            for gi, group in enumerate(groups):
                sid = f'combined_group{gi+1}'
                label = f"合并Skill(组{gi+1})"
                yield _sse({'type': 'step_start', 'question_idx': q_idx,
                            'step_id': sid, 'step_label': label,
                            'system_prompt': COMBINED_SKILL_SYSTEM[:300],
                            'user_prompt': f'组{gi+1}: {len(group)}篇文章'})

            if not groups:
                _log(f"[stability] row={row_num} 分组结果为空，跳过并行抽取")
                all_cc_pairs = []
                all_scenes = []
                all_relations = []
                all_time_constraints = []
            else:
                with ThreadPoolExecutor(max_workers=min(thread_count, len(groups))) as executor:
                    for gi, group in enumerate(groups):
                        executor.submit(_run_combined, gi, group)

                    all_cc_pairs = []
                    all_scenes = []
                    all_relations = []
                    all_time_constraints = []
                    completed = 0

                    while completed < len(groups):
                        sid, label, resp, parsed, error, group_art_ids = task_queue.get()
                        completed += 1
                        if error:
                            _log(f"[stability] row={row_num} {label}失败: {error}")
                            yield _sse({'type': 'step_error', 'question_idx': q_idx,
                                        'step_id': sid, 'step_label': label, 'error': error})
                        else:
                            _log(f"[stability] row={row_num} {label}完成，原始响应:\n{resp}")
                            if parsed and isinstance(parsed, dict):
                                cc = parsed.get('condition_conclusion_pairs', [])
                                scenes = parsed.get('policy_scenes', [])
                                rels = parsed.get('concept_relations', [])
                                tcs = parsed.get('time_constraints', [])
                                # 概念关系字段名映射
                                for item in rels:
                                    if isinstance(item, dict):
                                        item.setdefault('entity_a', item.pop('concept_A', ''))
                                        item.setdefault('entity_b', item.pop('concept_B', ''))
                                        item.setdefault('relation_type', item.pop('relation', ''))
                                # 规范化article_ids
                                for item in cc + tcs + rels:
                                    if isinstance(item, dict):
                                        aid = item.pop('article_id', None)
                                        if aid and 'article_ids' not in item:
                                            item['article_ids'] = [aid] if isinstance(aid, str) else aid
                                        # 如果仍无article_ids，用本组的文章ID回填
                                        if not item.get('article_ids'):
                                            item['article_ids'] = list(group_art_ids)
                                all_cc_pairs.extend(cc if isinstance(cc, list) else [])
                                all_scenes.extend(scenes if isinstance(scenes, list) else [])
                                all_relations.extend(rels if isinstance(rels, list) else [])
                                all_time_constraints.extend(tcs if isinstance(tcs, list) else [])
                                _log(f"[stability] row={row_num} {label} → 条件-结论={len(cc)}, "
                                     f"场景={len(scenes)}, 断言={len(rels)}, 时间={len(tcs)}")
                            yield _sse({'type': 'step_complete', 'question_idx': q_idx,
                                        'step_id': sid, 'step_label': label,
                                        'response': resp, 'parsed': parsed})

            # 2c: 跨组合并去重
            _log(f"\n{'='*60}\n[stability] ----------------Step 2c: 跨组合并去重---------------------\n{'='*60}")
            _log(f"[stability] row={row_num} 合并前（条件-结论={len(all_cc_pairs)}, "
                 f"场景={len(all_scenes)}, 断言={len(all_relations)}, "
                 f"时间约束={len(all_time_constraints)}）")

            products['condition_pairs'] = _merge_condition_pairs(all_cc_pairs)
            products['scene_enum'] = _dedupe_strings(all_scenes)
            products['assertions_raw'] = all_relations
            products['time_constraints'] = _merge_time_constraints(all_time_constraints)

            _log(f"[stability] row={row_num} 合并结果: 条件-结论={len(products['condition_pairs'])}条, "
                 f"场景={len(products['scene_enum'])}个, 断言={len(products['assertions_raw'])}条, "
                 f"时间约束={len(products['time_constraints'])}条")

            # ---- Step 2d: 三层过滤 ----
            _log(f"\n{'='*60}\n[stability] ----------------Step 2d: 三层过滤（相关性+置信度+冲突消解）---------------------\n{'='*60}")
            yield _sse({'type': 'step_start', 'question_idx': q_idx,
                        'step_id': 'filtering', 'step_label': '三层过滤(相关性+置信度+冲突消解)',
                        'system_prompt': '(纯代码规则过滤，不调用LLM)' if FILTER_ENABLED else '(已禁用)',
                        'user_prompt': question[:200]})
            if FILTER_ENABLED:
                try:
                    filtered = _filter_skills_outputs(question, products)
                    products['filtered_condition_pairs'] = filtered['condition_pairs']
                    products['filtered_scene_enum'] = filtered['policy_scenes']
                    products['filtered_assertions'] = filtered['concept_relations']
                    products['filtered_time_constraints'] = filtered['time_constraints']
                    _log(f"[stability] row={row_num} 三层过滤完成: "
                         f"条件-结论={len(filtered['condition_pairs'])}条, "
                         f"场景={len(filtered['policy_scenes'])}个, "
                         f"断言={len(filtered['concept_relations'])}条, "
                         f"时间={len(filtered['time_constraints'])}条")
                except Exception as e:
                    _log(f"[stability] row={row_num} 三层过滤失败，使用合并后数据: {e}")
                    products['filtered_condition_pairs'] = products['condition_pairs']
                    products['filtered_scene_enum'] = products['scene_enum']
                    products['filtered_assertions'] = products['assertions_raw']
                    products['filtered_time_constraints'] = products['time_constraints']
            else:
                # 过滤未启用，标记占位符
                _log(f"[stability] row={row_num} 三层过滤已禁用(FILTER_ENABLED=False)")
                products['filtered_condition_pairs'] = '阈值过滤逻辑未启用'
                products['filtered_scene_enum'] = '阈值过滤逻辑未启用'
                products['filtered_assertions'] = '阈值过滤逻辑未启用'
                products['filtered_time_constraints'] = '阈值过滤逻辑未启用'
            yield _sse({'type': 'step_complete', 'question_idx': q_idx,
                        'step_id': 'filtering', 'step_label': '三层过滤(相关性+置信度+冲突消解)',
                        'response': f"过滤后: CC={len(products['filtered_condition_pairs']) if isinstance(products['filtered_condition_pairs'], list) else '未启用'}, "
                                    f"场景={len(products['filtered_scene_enum']) if isinstance(products['filtered_scene_enum'], list) else '未启用'}, "
                                    f"断言={len(products['filtered_assertions']) if isinstance(products['filtered_assertions'], list) else '未启用'}, "
                                    f"时间={len(products['filtered_time_constraints']) if isinstance(products['filtered_time_constraints'], list) else '未启用'}"})

            # ---- Step 2e: 多维评分过滤 ----
            _log(f"\n{'='*60}\n[stability] ----------------Step 2e: 多维评分过滤（按权重排序截断）---------------------\n{'='*60}")
            yield _sse({'type': 'step_start', 'question_idx': q_idx,
                        'step_id': 'score_filtering', 'step_label': '多维评分过滤(按权重排序截断)',
                        'system_prompt': '(四维评分排序，不调用LLM)' if SCORE_FILTER_ENABLED else '(已禁用)',
                        'user_prompt': question[:200]})
            if SCORE_FILTER_ENABLED:
                try:
                    score_filtered = _score_filter_products(question, article_parts, products)
                    products['score_filtered_condition_pairs'] = score_filtered['condition_pairs']
                    products['score_filtered_scene_enum'] = score_filtered['policy_scenes']
                    products['score_filtered_assertions'] = score_filtered['concept_relations']
                    products['score_filtered_time_constraints'] = score_filtered['time_constraints']
                    _log(f"[stability] row={row_num} 多维评分过滤完成: "
                         f"条件-结论={len(score_filtered['condition_pairs'])}条, "
                         f"场景={len(score_filtered['policy_scenes'])}个, "
                         f"断言={len(score_filtered['concept_relations'])}条, "
                         f"时间={len(score_filtered['time_constraints'])}条")
                except Exception as e:
                    _log(f"[stability] row={row_num} 多维评分过滤失败，使用合并后数据: {e}")
                    products['score_filtered_condition_pairs'] = products['condition_pairs']
                    products['score_filtered_scene_enum'] = products['scene_enum']
                    products['score_filtered_assertions'] = products['assertions_raw']
                    products['score_filtered_time_constraints'] = products['time_constraints']
            else:
                _log(f"[stability] row={row_num} 多维评分过滤已禁用(SCORE_FILTER_ENABLED=False)")
                products['score_filtered_condition_pairs'] = '权重过滤逻辑未启用'
                products['score_filtered_scene_enum'] = '权重过滤逻辑未启用'
                products['score_filtered_assertions'] = '权重过滤逻辑未启用'
                products['score_filtered_time_constraints'] = '权重过滤逻辑未启用'
            yield _sse({'type': 'step_complete', 'question_idx': q_idx,
                        'step_id': 'score_filtering', 'step_label': '多维评分过滤(按权重排序截断)',
                        'response': f"权重过滤后: CC={len(products['score_filtered_condition_pairs']) if isinstance(products['score_filtered_condition_pairs'], list) else '未启用'}, "
                                    f"场景={len(products['score_filtered_scene_enum']) if isinstance(products['score_filtered_scene_enum'], list) else '未启用'}, "
                                    f"断言={len(products['score_filtered_assertions']) if isinstance(products['score_filtered_assertions'], list) else '未启用'}, "
                                    f"时间={len(products['score_filtered_time_constraints']) if isinstance(products['score_filtered_time_constraints'], list) else '未启用'}"})

            # ---- Step 3: Validate assertions ----
            _log(f"\n{'='*60}\n[stability] ----------------Step 3: 断言校验 + 转自然语言---------------------\n{'='*60}")
            # 数据源优先级：权重过滤 > 阈值过滤 > 合并后
            if SCORE_FILTER_ENABLED and isinstance(products.get('score_filtered_condition_pairs'), list):
                cc_for_prompt = products['score_filtered_condition_pairs']
                scene_for_prompt = products['score_filtered_scene_enum']
                assertion_for_validate = products['score_filtered_assertions']
                tc_for_prompt = products['score_filtered_time_constraints']
            elif FILTER_ENABLED:
                cc_for_prompt = products['filtered_condition_pairs']
                scene_for_prompt = products['filtered_scene_enum']
                assertion_for_validate = products['filtered_assertions']
                tc_for_prompt = products['filtered_time_constraints']
            else:
                cc_for_prompt = products['condition_pairs']
                scene_for_prompt = products['scene_enum']
                assertion_for_validate = products['assertions_raw']
                tc_for_prompt = products['time_constraints']

            raw_cc_count = len(cc_for_prompt)
            raw_scene_count = len(scene_for_prompt)
            raw_tc_count = len(tc_for_prompt)

            raw_count = len(assertion_for_validate)
            cleaned = _validate_assertions(assertion_for_validate)
            products['assertions_cleaned'] = cleaned
            constraint_texts = _convert_constraints_to_text(cleaned)
            products['constraint_texts'] = constraint_texts

            _log(f"[stability] row={row_num} 断言校验: {raw_count}→{len(cleaned)}")
            _log(f"[stability] row={row_num} 条件-结论对(合并后): {raw_cc_count}条")
            _log(f"[stability] row={row_num} 政策场景(去重后): {raw_scene_count}个")
            _log(f"[stability] row={row_num} 时间约束(合并后): {raw_tc_count}条")
            yield _sse({'type': 'validation', 'question_idx': q_idx,
                        'original_count': raw_count, 'cleaned_count': len(cleaned),
                        'removed_count': raw_count - len(cleaned),
                        'details': f"原始{raw_count}条 → 清洗后{len(cleaned)}条"})

            # ---- Step 4: Assemble final prompt ----
            _log(f"\n{'='*60}\n[stability] ----------------Step 4: 组装最终提示词（Prompt.md模板）---------------------\n{'='*60}")
            _log(f"[stability] row={row_num} 条件-结论对={raw_cc_count}条, 政策场景={raw_scene_count}个, 时间约束={raw_tc_count}条, 概念约束={len(constraint_texts)}条")
            final_prompt = _assemble_final_prompt(
                question,
                cc_for_prompt,
                scene_for_prompt,
                constraint_texts,
                tc_for_prompt
            )
            products['final_prompt'] = final_prompt
            _log(f"[stability] row={row_num} 最终提示词组装完成（长度={len(final_prompt)}）")
            yield _sse({'type': 'final_prompt', 'question_idx': q_idx, 'prompt': final_prompt})

            # ---- Step 5: Generate final answer ----
            _log(f"\n{'='*60}\n[stability] ----------------Step 5: 生成最终回答（当前已禁用）---------------------\n{'='*60}")
            yield _sse({'type': 'step_start', 'question_idx': q_idx,
                        'step_id': 'final_answer', 'step_label': '生成最终回答',
                        'system_prompt': FINAL_ANSWER_SYSTEM, 'user_prompt': final_prompt})
            # # 暂时注释掉最终回答生成，先输出组装后的提示词
            # try:
            #     final_answer = _call_llm(FINAL_ANSWER_SYSTEM, final_prompt,
            #                              temperature=0.3, max_tokens=4000, timeout=600)
            #     products['final_answer'] = final_answer
            #     logger.info(f"[stability] row={row_num} 最终回答完成（长度={len(final_answer)}）")
            #     yield _sse({'type': 'step_complete', 'question_idx': q_idx,
            #                 'step_id': 'final_answer', 'step_label': '生成最终回答',
            #                 'response': final_answer})
            # except Exception as e:
            #     fallback = "系统暂无法回答此问题，请稍后重试或咨询税务专员。"
            #     products['final_answer'] = fallback
            #     logger.error(f"[stability] row={row_num} 最终回答生成失败: {e}")
            #     yield _sse({'type': 'step_error', 'question_idx': q_idx,
            #                 'step_id': 'final_answer', 'step_label': '生成最终回答',
            #                 'error': str(e)})
            products['final_answer'] = "（最终回答生成已禁用，请查看上方最终组装的提示词）"
            yield _sse({'type': 'step_complete', 'question_idx': q_idx,
                        'step_id': 'final_answer', 'step_label': '生成最终回答',
                        'response': products['final_answer']})

            yield _sse({'type': 'question_complete', 'question_idx': q_idx,
                        'answer': products['final_answer']})

            all_results.append({
                'row': row_num,
                'question': question,
                'prompt': kb_prompt,
                'articles_text': products['articles_text'],
                'raw_condition_pairs': all_cc_pairs,
                'condition_pairs': products['condition_pairs'],
                'filtered_condition_pairs': products.get('filtered_condition_pairs', []),
                'score_filtered_condition_pairs': products.get('score_filtered_condition_pairs', []),
                'raw_scene_enum': all_scenes,
                'scene_enum': products['scene_enum'],
                'filtered_scene_enum': products.get('filtered_scene_enum', []),
                'score_filtered_scene_enum': products.get('score_filtered_scene_enum', []),
                'raw_assertions': all_relations,
                'assertions_cleaned': products['assertions_cleaned'],
                'filtered_assertions': products.get('filtered_assertions', []),
                'score_filtered_assertions': products.get('score_filtered_assertions', []),
                'constraint_texts': products['constraint_texts'],
                'raw_time_constraints': all_time_constraints,
                'time_constraints': products['time_constraints'],
                'filtered_time_constraints': products.get('filtered_time_constraints', []),
                'score_filtered_time_constraints': products.get('score_filtered_time_constraints', []),
                'final_prompt': products['final_prompt'],
                'final_answer': products['final_answer']
            })

        # Save results
        try:
            stability_output_dir = os.path.join(OUTPUT_FOLDER, '回答稳定性结果')
            os.makedirs(stability_output_dir, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"回答稳定性结果_{ts}.xlsx"
            output_path = os.path.join(stability_output_dir, output_filename)
            _save_results(all_results, output_path)
            _log(f"[stability] 结果已保存: {output_path}")
        except Exception as e:
            _log(f"[stability] 保存结果失败: {e}")
            output_filename = ""

        _log(f"[stability] ===== 全部完成 ({total}题) =====")
        yield _sse({'type': 'complete', 'total_questions': total,
                    'output_filename': f'回答稳定性结果/{output_filename}'})

    return Response(generate(), mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


def _save_results(results, output_path):
    """Save results to Excel with pre/post processing comparison"""
    wb = Workbook()
    ws = wb.active
    ws.title = '回答稳定性结果'

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        ('问题', 50),
        ('提示词(B列)', 40),
        ('文章内容', 80),
        ('1.条件-结论对(处理前)', 60),
        ('2.条件-结论对(处理后)', 60),
        ('3.条件-结论对(过滤后)', 60),
        ('4.条件-结论对(按权重过滤)', 60),
        ('1.政策场景(处理前)', 30),
        ('2.政策场景(处理后)', 30),
        ('3.政策场景(过滤后)', 30),
        ('4.政策场景(按权重过滤)', 30),
        ('1.概念断言(处理前)', 50),
        ('2.概念断言(过滤后)', 50),
        ('3.概念断言(按权重过滤)', 50),
        ('4.概念断言(清洗后)', 50),
        ('5.概念约束文本', 40),
        ('1.时间约束(处理前)', 40),
        ('2.时间约束(处理后)', 40),
        ('3.时间约束(过滤后)', 40),
        ('4.时间约束(按权重过滤)', 40),
        ('最终提示词', 80),
        ('最终回答', 80)
    ]

    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, r in enumerate(results, 2):
        def _fmt_filtered(val):
            """格式化过滤后数据：字符串直接返回（未启用提示），列表用json.dumps"""
            if isinstance(val, str):
                return val
            return json.dumps(val, ensure_ascii=False, indent=2)

        values = [
            r['question'],                                                              # 1.问题
            r.get('prompt', ''),                                                        # 2.提示词(B列)
            r.get('articles_text', ''),                                                 # 3.文章内容
            json.dumps(r.get('raw_condition_pairs', []), ensure_ascii=False, indent=2), # 4.条件-结论对(处理前)
            json.dumps(r.get('condition_pairs', []), ensure_ascii=False, indent=2),     # 5.条件-结论对(处理后)
            _fmt_filtered(r.get('filtered_condition_pairs', [])),                       # 6.条件-结论对(过滤后)
            _fmt_filtered(r.get('score_filtered_condition_pairs', [])),                 # 7.条件-结论对(按权重过滤)
            json.dumps(r.get('raw_scene_enum', []), ensure_ascii=False),                # 8.政策场景(处理前)
            json.dumps(r.get('scene_enum', []), ensure_ascii=False),                    # 9.政策场景(处理后)
            _fmt_filtered(r.get('filtered_scene_enum', [])),                            # 10.政策场景(过滤后)
            _fmt_filtered(r.get('score_filtered_scene_enum', [])),                      # 11.政策场景(按权重过滤)
            json.dumps(r.get('raw_assertions', []), ensure_ascii=False, indent=2),      # 12.概念断言(处理前)
            _fmt_filtered(r.get('filtered_assertions', [])),                            # 13.概念断言(过滤后)
            _fmt_filtered(r.get('score_filtered_assertions', [])),                      # 14.概念断言(按权重过滤)
            json.dumps(r.get('assertions_cleaned', []), ensure_ascii=False, indent=2),  # 15.概念断言(清洗后)
            '\n'.join(r.get('constraint_texts', [])),                                   # 16.概念约束文本
            json.dumps(r.get('raw_time_constraints', []), ensure_ascii=False, indent=2),# 17.时间约束(处理前)
            json.dumps(r.get('time_constraints', []), ensure_ascii=False, indent=2),    # 18.时间约束(处理后)
            _fmt_filtered(r.get('filtered_time_constraints', [])),                      # 19.时间约束(过滤后)
            _fmt_filtered(r.get('score_filtered_time_constraints', [])),                # 20.时间约束(按权重过滤)
            r.get('final_prompt', ''),                                                  # 21.最终提示词
            r.get('final_answer', '')                                                   # 22.最终回答
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border

    wb.save(output_path)
