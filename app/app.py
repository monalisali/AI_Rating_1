#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Web应用 - 提供Excel上传页面并调用API处理
支持AI回答和语义对比打分
"""
"""
model参数，如：
百炼模型智谱：glm-5.1
GTS模型智谱：saas.glm-5.1


1. GTS模型配置：
url: https://genai-sharedservice-uat.cn.asia.pwcinternal.com
key: sk-JLfi9OVwSLESgKy0Xw_N2w
model:
-saas.glm-5.1
-saas.qwen3.5-plus

2. 智谱官方
url: https://open.bigmodel.cn/api/paas/v4
key: f5d8c53a2872430fb5de64c5c690bbf9.a3WyEO2pOajsX404
model：
-glm-5.1

3. 加油站
url: http://ai.tech.tax.asia.pwcinternal.com:3002
key: sk-wKJ3gwHNsyZL3oZDZrOMYlQbLbCxY8QpLDxRYMqCGp9Ms4fz
model:
- bedrock.anthropic.claude-opus-4-7 (很贵)

"""


import os
import re
import json
import ssl
import logging
import urllib.request
import queue
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side

# 项目根目录（app.py在app/子目录下，根目录是上一级）
ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 日志配置：同时输出到控制台和文件
LOG_DIR = os.path.join(ROOT_DIR, 'logs')
os.makedirs(LOG_DIR, exist_ok=True)
LOG_FILE = os.path.join(LOG_DIR, f"optimize_{datetime.now().strftime('%Y%m%d')}.log")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

_log_fmt = logging.Formatter('%(asctime)s %(levelname)s %(message)s')

_file_handler = logging.FileHandler(LOG_FILE, encoding='utf-8')
_file_handler.setLevel(logging.INFO)
_file_handler.setFormatter(_log_fmt)
logger.addHandler(_file_handler)

_console_handler = logging.StreamHandler()
_console_handler.setLevel(logging.INFO)
_console_handler.setFormatter(_log_fmt)
logger.addHandler(_console_handler)

app = Flask(__name__, template_folder=os.path.join(ROOT_DIR, 'templates'))
app.config['UPLOAD_FOLDER'] = os.path.join(ROOT_DIR, 'uploads')
app.config['OUTPUT_FOLDER'] = os.path.join(ROOT_DIR, 'outputs')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 最大16MB
app.config['JSON_AS_ASCII'] = False

# 确保文件夹存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# Prompt.md 和 config.json 在项目根目录
PROMPT_FILE = os.path.join(ROOT_DIR, 'Prompt.md')
CONFIG_FILE = os.path.join(ROOT_DIR, 'config.json')


def get_config(key, default=None):
    """从config.json读取配置（每次调用重新读取，支持运行时修改）"""
    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            cfg = json.load(f)
        return cfg.get(key, default)
    except Exception:
        return default


def load_system_prompt():
    """加载 Prompt.md 作为知识库API的自定义系统提示词"""
    if os.path.exists(PROMPT_FILE):
        with open(PROMPT_FILE, 'r', encoding='utf-8') as f:
            return f.read().strip()
    logger.warning(f"Prompt.md 不存在: {PROMPT_FILE}")
    return ""


# 评分模型配置 - 优先从config.json读取，环境变量兜底
def _get_scoring_config():
    url = get_config('scoring_api_url', '')
    if not url:
        url = os.environ.get('ANTHROPIC_BASE_URL', 'http://ai.tech.tax.asia.pwcinternal.com:3002') + '/v1/chat/completions'
    key = get_config('scoring_api_key', '')
    if not key:
        key = os.environ.get('ANTHROPIC_AUTH_TOKEN', '')
    model = get_config('scoring_model', '')
    if not model:
        model = os.environ.get('ANTHROPIC_MODEL', 'glm-coding-5-8')
    return url, key, model

SCORING_API_URL, SCORING_API_KEY, SCORING_MODEL = _get_scoring_config()
logger.info(f"[config] 评分API: {SCORING_API_URL}, 模型: {SCORING_MODEL}, Key: {'已配置' if SCORING_API_KEY else '未配置'}")


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def request_api(message: str, session_id: str = "", custom_system_prompt: str = "", max_retries: int = 3, timeout: int = 600) -> tuple:
    """请求知识库API接口，线程级超时保护，自动重试"""
    import http.client
    import threading

    url = 'https://ai.tech.tax.asia.pwcinternal.com:5007/api/chat-stream'
    payload = {
        'message': message,
        'session_id': session_id,
        'model': get_config('model', 'glm-5.1')
    }
    if custom_system_prompt:
        payload['custom_system_prompt'] = custom_system_prompt
    llm_api_key = get_config('llm_api_key', '')
    llm_base_url = get_config('llm_base_url', '')
    if llm_api_key:
        payload['llm_api_key'] = llm_api_key
    if llm_base_url:
        payload['llm_base_url'] = llm_base_url
    data = json.dumps(payload).encode('utf-8')

    req = urllib.request.Request(url, data=data, headers={
        'Content-Type': 'application/json',
        'Accept': 'text/event-stream'
    })

    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    for attempt in range(1, max_retries + 1):
        result_holder = [None, None]  # [result_tuple, error]

        def _call():
            try:
                with urllib.request.urlopen(req, context=ctx, timeout=timeout) as response:
                    returned_session_id = response.headers.get('X-Session-Id', '')
                    chunks = []
                    while True:
                        chunk = response.read(8192)
                        if not chunk:
                            break
                        chunks.append(chunk)
                    result = b''.join(chunks).decode('utf-8')
                result_holder[0] = (result, returned_session_id)
            except Exception as e:
                result_holder[1] = e

        t = threading.Thread(target=_call, daemon=True)
        t.start()
        t.join(timeout=timeout + 60)

        if t.is_alive():
            logger.warning(f"[request_api] 第{attempt}次请求线程超时（>{timeout + 60}秒），强制跳过")
            if attempt >= max_retries:
                raise TimeoutError(f"5007 API调用超时（>{timeout + 60}秒），已重试{max_retries}次")
            continue

        if result_holder[1] is not None:
            err = result_holder[1]
            if isinstance(err, (http.client.IncompleteRead, ConnectionError, TimeoutError)):
                logger.warning(f"[request_api] 第{attempt}次请求失败({type(err).__name__}: {err})，{'重试中...' if attempt < max_retries else '已达最大重试次数'}")
                if attempt >= max_retries:
                    raise
            else:
                raise err

        if result_holder[0] is not None:
            return result_holder[0]

    raise RuntimeError("request_api: 未获取到结果")


def parse_response(api_response: str) -> dict:
    """解析SSE流式API响应"""
    contents = []
    error_msg = None
    for line in api_response.strip().split('\n'):
        line = line.strip()
        if line.startswith('data:'):
            try:
                data = json.loads(line[5:].strip())
                if data.get('type') == 'content':
                    contents.append(data.get('content', ''))
                elif data.get('type') == 'error':
                    error_msg = data.get('content', str(data))
                    logger.error(f"[parse_response] API返回错误: {error_msg}")
            except json.JSONDecodeError:
                pass

    full = ''.join(contents)

    # 如果有错误且没有任何内容，抛异常让调用方重试
    if error_msg and not full.strip():
        raise RuntimeError(f"5007 API错误: {error_msg}")

    return {'full_content': full}


def is_confirmation_question(content: str) -> bool:
    """判断内容是否是确认问句"""
    keywords = [
        '请问以上关键词是否需要调整或补充', '确认后我将开始',
        '请问这样理解是否正确', '您是否有需要补充或调整',
        '是否需要调整', '请确认', '确认后', '以上关键词是否准确',
        '是否需要添加', '请问以上', '确认后开始', '是否准确',
        '请告知', '是否继续', '我将开始', '以上内容是否',
        '是否合适', '希望调整', '是否需要修改', '是否满意',
        '是否同意', '请告诉我您的修改', '需要调整请告诉我',
    ]
    return any(kw in content for kw in keywords)


def is_incomplete_answer(content: str) -> bool:
    """判断内容是否是不完整的中间步骤（没有核心发现且较短）"""
    if '核心发现' in content:
        return False
    # 内容过短说明还没拿到完整答案
    return len(content.strip()) < 200


def chat_with_confirmation(question: str, max_rounds: int = 8, system_prompt: str = "") -> str:
    """执行多轮对话，自动处理确认和等待完整答案，API错误时重试"""
    session_id = ""
    current_message = question
    content = ""

    for round_i in range(max_rounds):
        try:
            api_response, session_id = request_api(current_message, session_id, custom_system_prompt=system_prompt)
            content = parse_response(api_response)['full_content']
            if is_confirmation_question(content):
                current_message = "同意，请使用这些关键词进行搜索，不需要调整。"
            elif is_incomplete_answer(content):
                current_message = "继续"
            else:
                return content
        except RuntimeError as e:
            logger.warning(f"[chat_with_confirmation] 第{round_i+1}轮API错误: {e}，重置session重试...")
            session_id = ""
            current_message = question

    return content


def request_scoring_api(prompt: str, timeout: int = 300, max_retries: int = 3) -> str:
    """请求内网AI评分API，带线程超时保护防止卡死，遇到429自动重试"""
    import threading
    import time

    for retry in range(1, max_retries + 1):
        result_holder = [None, None]  # [result, error]

        def _call():
            try:
                url, key, model = _get_scoring_config()
                data = json.dumps({
                    'model': model,
                    'messages': [{'role': 'user', 'content': prompt}]
                }).encode('utf-8')

                req = urllib.request.Request(url, data=data, headers={
                    'Content-Type': 'application/json',
                    'Authorization': f'Bearer {key}'
                })

                with urllib.request.urlopen(req, timeout=timeout) as response:
                    resp = json.loads(response.read().decode('utf-8'))
                    if 'error' in resp:
                        err_msg = resp['error'].get('message', json.dumps(resp['error'], ensure_ascii=False)) if isinstance(resp['error'], dict) else str(resp['error'])
                        result_holder[1] = ValueError(f"评分API返回错误: {err_msg}")
                        return
                    if 'choices' not in resp or not resp['choices']:
                        result_holder[1] = ValueError(f"API返回格式异常: {json.dumps(resp, ensure_ascii=False)[:300]}")
                        return
                    result_holder[0] = resp['choices'][0]['message']['content']
            except Exception as e:
                result_holder[1] = e

        t = threading.Thread(target=_call, daemon=True)
        t.start()
        t.join(timeout=timeout + 30)

        if t.is_alive():
            raise TimeoutError(f"评分API调用超过{timeout + 30}秒，已强制终止")
        if result_holder[1] is not None:
            err = result_holder[1]
            is_429 = '429' in str(err)
            if is_429 and retry < max_retries:
                wait = retry * 10
                logger.warning(f"[request_scoring_api] 429限流，等待{wait}秒后重试（{retry}/{max_retries}）")
                time.sleep(wait)
                continue
            raise err
        return result_holder[0]


# 默认评分提示词模板（可通过前端覆盖）
DEFAULT_SCORING_PROMPT = """请作为一名专业的税务领域评估专家，对以下AI回答进行评分。

【问题】
{question}

【参考答案】
{reference_answer}

【AI回答】
{ai_answer}

请从以下3个维度进行评分，并以JSON格式返回结果：

1. 答案准确性（60分）：
   - 将AI回答中"核心发现"部分的每一条要点与参考答案进行逐条语义对比，重点判断核心结论是否准确、有无遗漏或偏差。"核心发现"是回答的精华摘要，其准确性权重最高。
   - 对回答整体内容与参考答案进行全面语义对比，评估事实、数据、税率、政策适用等方面是否准确。
   - 评估回答是否针对问题所涉及的具体地区（如中国大陆、香港、特定省市等）的税收政策和规定进行了准确回答。如果问题涉及特定地区，回答是否正确区分了该地区的特殊规定；如果回答了不相关的地区政策应扣分。若问题不涉及地区性差异则不扣分。
   - 重点检查"核心发现"部分的结论与回答后续展开的详细分析、案例、数据之间是否存在矛盾或冲突。例如：核心发现说适用税率A，但详细分析中又说适用税率B；或核心发现说某政策适用，但详细分析中列出了不满足的条件。存在此类矛盾应扣分。

   【严重偏差判定——符合以下任一情形，准确性直接打0分】：
   a) 遗漏关键税务处理步骤：参考答案中的核心税务调整步骤（如视同销售后的费用扣除基数联动调整、特定申报表行次填报、跨表勾稽关系等）在AI回答中完全未提及，导致整条计算链条断裂。
   b) 核心计算基数错误：AI回答中用于关键计算的基础数据（如收入额、扣除限额的计税基础、视同销售的公允价值等）与参考答案不一致，导致后续计算结果全部偏误。
   c) 引入题干明确排除的无关内容且挤占关键篇幅：题干已限定范围（如"仅企业所得税""不考虑增值税"），AI回答却大篇幅讨论题干明确排除的税种或无关政策，且该无关内容导致关键要点被遗漏或论述不足。
   d) 核心要点遗漏过半：参考答案包含多个独立的核心结论或处理步骤，AI回答遗漏其中超过半数。
   - 以上方面均准确无误、核心发现与参考答案完全一致且无矛盾给60分。
   - 当准确性评分为0分时，法条援引度和总结完整度也全部直接置为0分。

2. 法条援引度（20分）：
   - 检查AI回答中引用的法规条款是否正确、是否与参考答案中提及的法条一致。
   - 检查回答末尾或正文中的引用法规列表/参考资料，判断这些法规是否在回答正文中被实际引用和讨论。
   - 检查回答正文中提到的法规是否都在引用列表中列出，确保引用列表与正文引用完全对应。
   - 引用正确完整、列表与正文完全对应给20分。

3. 总结完整度（20分）：
   - 评估"核心发现"作为总结是否完整，是否涵盖了问题的所有关键方面，是否遗漏了参考答案中的重要要点。
   - 检查回答是否对问题中隐含的子问题或特殊情况都给出了回应。
   - 总结全面、覆盖完整给20分。

请严格按照以下JSON格式返回（不要添加任何其他文字）：
{{
    "accuracy_score": 分数,
    "accuracy_reason": "评分说明",
    "citation_score": 分数,
    "citation_reason": "评分说明",
    "summary_score": 分数,
    "summary_reason": "评分说明"
}}"""


def optimize_prompt(current_prompt: str, results_with_scores: list, attempt: int = 1, optimize_template: str = None, context_info: str = "") -> str:
    """
    根据评分结果，定向优化系统提示词

    Args:
        current_prompt: 当前使用的系统提示词
        results_with_scores: 列表，每项包含 question, answer, reference_answer, scores
        attempt: 第几次尝试
        optimize_template: 外层优化过的优化指令模板（None则用默认）

    Returns:
        优化后的新系统提示词
    """
    logger.info("[optimize_prompt] 开始优化提示词...")

    # 构建评分详情（不截取）
    details = []
    low_dims = {'accuracy': False, 'citation': False, 'summary': False}
    for i, r in enumerate(results_with_scores):
        scores = r.get('scores', {})
        detail = f"问题{i+1}: {r['question']}\n"
        detail += f"AI回答: {r['answer']}\n"
        if r.get('reference_answer'):
            detail += f"参考答案: {r['reference_answer']}\n"
        if scores and scores.get('success'):
            detail += f"总分: {scores['total_score']}/100\n"
            detail += f"准确性: {scores['accuracy_score']}/60 - {scores['accuracy_reason']}\n"
            detail += f"法条援引: {scores['citation_score']}/20 - {scores['citation_reason']}\n"
            detail += f"总结完整度: {scores['summary_score']}/20 - {scores['summary_reason']}\n"
            if scores['accuracy_score'] < 48:
                low_dims['accuracy'] = True
            if scores['citation_score'] < 16:
                low_dims['citation'] = True
            if scores['summary_score'] < 16:
                low_dims['summary'] = True
        else:
            detail += "评分失败\n"
        details.append(detail)

    # 根据低分维度生成定向优化指令
    focus_parts = []
    if low_dims['accuracy']:
        focus_parts.append(
            "- 【准确性偏低】重点优化搜索策略（步骤4和4.6），确保AI搜索更全面、不遗漏关键法规，"
            "加强法规内容与问题的关联分析，确保核心结论准确。不要修改与搜索无关的部分。"
        )
    if low_dims['citation']:
        focus_parts.append(
            "- 【法条援引偏低】重点优化引用列表与校验部分（步骤6），确保AI在回答中实际引用和讨论每条法规，"
            "引用列表与正文引用完全对应。不要修改与引用无关的部分。"
        )
    if low_dims['summary']:
        focus_parts.append(
            "- 【总结完整度偏低】重点优化核心发现和总结的生成要求（步骤5），确保AI的总结覆盖问题的所有关键方面，"
            "不遗漏重要要点。不要修改与总结无关的部分。"
        )

    focus_text = '\n'.join(focus_parts) if focus_parts else "- 整体评分偏低，请全面检查并优化系统提示词中的薄弱环节。"
    details_text = '\n'.join(details)

    # 使用外层优化的模板或默认模板
    if optimize_template:
        optimize_instruction = optimize_template.format(
            current_prompt=current_prompt,
            details=details_text,
            focus_text=focus_text
        )
    else:
        optimize_instruction = f"""你是一个专业的提示词优化专家。当前有一个用于税务法规知识库AI助手的系统提示词，但使用该提示词后AI回答的评分不够理想。

请分析评分数据中低分维度的具体原因，**对现有规则进行精准修改**，而非堆砌新规则。

【当前系统提示词】
{current_prompt}

【各题评分详情】
{details_text}

【需要优化的维度】
{focus_text}

【修改原则】
1. **精准修改**：根据评分详情中的扣分原因，找到导致该问题的具体规则并针对性修改
2. **允许补充和新增**：可以在现有段落内补充约束，也可以新增必要的步骤或章节，但新增内容要精炼、直接解决扣分问题，不要堆砌冗长的规则列表
3. **未修改部分必须原样保留**：与低分维度无关的所有内容不得有任何改动
4. 不要输出任何解释说明，只输出新的系统提示词本身
"""

    try:
        new_prompt = request_scoring_api(optimize_instruction)
        max_relative_ratio = get_config('max_prompt_ratio', 1.5)
        max_relative_len = int(len(current_prompt) * max_relative_ratio)
        max_absolute_len = get_config('max_prompt_length', 0)
        # 取相对限制和绝对限制中较小的值
        max_len = max_relative_len
        if max_absolute_len > 0:
            max_len = min(max_relative_len, max_absolute_len)
        # 判断是否会被丢弃
        discarded = len(new_prompt) > max_len
        status = "已丢弃" if discarded else "已采用"
        reason = ""
        if discarded:
            reasons = []
            if len(new_prompt) > max_relative_len:
                reasons.append(f"长度{len(new_prompt)}超过相对上限{max_relative_len}（原始{len(current_prompt)}的150%）")
            if max_absolute_len > 0 and len(new_prompt) > max_absolute_len:
                reasons.append(f"长度{len(new_prompt)}超过绝对上限{max_absolute_len}（配置项max_prompt_length）")
            reason = f"\n丢弃原因: {'; '.join(reasons)}"
        # 保存到文件（无论是否被丢弃）
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        prompt_file = os.path.join(LOG_DIR, f"optimized_prompt_{ts[:8]}.txt")
        with open(prompt_file, 'a', encoding='utf-8') as f:
            f.write(f"\n{'='*60}\n")
            f.write(f"===== 第{attempt}次优化生成的提示词 ({ts}) [{status}] =====\n")
            if context_info:
                f.write(f"{context_info}\n")
            f.write(f"状态: {status} | 原始长度={len(current_prompt)} → 新长度={len(new_prompt)} | 上限={max_len}\n")
            if reason:
                f.write(f"{reason}\n")
            f.write(f"{'='*60}\n\n")
            f.write(new_prompt)
            f.write("\n")
        if discarded:
            logger.warning(f"[optimize_prompt] 优化后提示词过长（{len(new_prompt)} > {max_len}），已记录但丢弃，保留原始提示词")
            return current_prompt
        logger.info(f"[optimize_prompt] 优化完成，原始长度={len(current_prompt)}, 新长度={len(new_prompt)}，已保存到: {prompt_file}")
        return new_prompt
    except Exception as e:
        logger.error(f"提示词优化失败: {e}")
        return current_prompt


def _build_score_details(results_with_scores):
    """构建评分详情文本（不截取，用于内层优化）"""
    details = []
    for i, r in enumerate(results_with_scores):
        scores = r.get('scores', {})
        detail = f"问题{i+1}: {r['question']}\n"
        detail += f"AI回答: {r['answer']}\n"
        if r.get('reference_answer'):
            detail += f"参考答案: {r['reference_answer']}\n"
        if scores and scores.get('success'):
            detail += f"总分: {scores['total_score']}/100\n"
            detail += f"准确性: {scores['accuracy_score']}/60 - {scores['accuracy_reason']}\n"
            detail += f"法条援引: {scores['citation_score']}/20 - {scores['citation_reason']}\n"
            detail += f"总结完整度: {scores['summary_score']}/20 - {scores['summary_reason']}\n"
        else:
            detail += "评分失败\n"
        details.append(detail)
    return '\n'.join(details)


def _build_focus_text(results_with_scores):
    """根据评分结果判断低分维度"""
    low_dims = {'accuracy': False, 'citation': False, 'summary': False}
    for r in results_with_scores:
        scores = r.get('scores', {})
        if scores and scores.get('success'):
            if scores['accuracy_score'] < 48:
                low_dims['accuracy'] = True
            if scores['citation_score'] < 16:
                low_dims['citation'] = True
            if scores['summary_score'] < 16:
                low_dims['summary'] = True

    focus_parts = []
    if low_dims['accuracy']:
        focus_parts.append("- 【准确性偏低】重点优化搜索策略（步骤4和4.6），确保AI搜索更全面、不遗漏关键法规，"
                           "加强法规内容与问题的关联分析，确保核心结论准确。")
    if low_dims['citation']:
        focus_parts.append("- 【法条援引偏低】重点优化引用列表与校验部分（步骤6），确保AI在回答中实际引用和讨论每条法规，"
                           "引用列表与正文引用完全对应。")
    if low_dims['summary']:
        focus_parts.append("- 【总结完整度偏低】重点优化核心发现和总结的生成要求（步骤5），确保AI的总结覆盖问题的所有关键方面，"
                           "不遗漏重要要点。")
    return '\n'.join(focus_parts) if focus_parts else "- 整体评分偏低，请全面检查并优化系统提示词中的薄弱环节。"


def optimize_optimization_method(optimize_template, all_attempt_logs):
    """
    外层循环：优化"优化方法"本身
    分析多轮优化仍低分的原因，让AI生成新的优化指令模板
    """
    logger.info("[optimize_method] 开始优化优化方法...")

    # 每次尝试的评分详情（AI回答和参考答案截取前500字，防止payload过大）
    TRUNCATE_LEN = 500
    all_details = []
    for log_entry in all_attempt_logs:
        detail = f"第{log_entry['attempt']}次尝试: 平均总分={log_entry['avg_total']:.1f}\n"
        for i, r in enumerate(log_entry['results']):
            scores = r.get('scores', {})
            detail += f"问题{i+1}: {r['question']}\n"
            detail += f"AI回答: {r['answer'][:TRUNCATE_LEN]}{'...(截断)' if len(r['answer']) > TRUNCATE_LEN else ''}\n"
            if r.get('reference_answer'):
                ref = r['reference_answer']
                detail += f"参考答案: {ref[:TRUNCATE_LEN]}{'...(截断)' if len(ref) > TRUNCATE_LEN else ''}\n"
            if scores and scores.get('success'):
                detail += f"总分: {scores['total_score']}/100\n"
                detail += f"准确性: {scores['accuracy_score']}/60 - {scores['accuracy_reason']}\n"
                detail += f"法条援引: {scores['citation_score']}/20 - {scores['citation_reason']}\n"
                detail += f"总结完整度: {scores['summary_score']}/20 - {scores['summary_reason']}\n"
            else:
                detail += "评分失败\n"
        all_details.append(detail)
    full_history = '\n'.join(all_details)

    last_results = all_attempt_logs[-1]['results']
    last_focus = _build_focus_text(last_results)
    score_threshold = get_config('score_threshold', 80)

    meta_instruction = f"""你是一个元优化专家。当前有一个用于优化"税务法规知识库AI助手系统提示词"的优化指令，但经过多轮使用后，评分始终无法达到{score_threshold}分。

请分析当前的优化指令存在什么问题，然后输出一个改进后的完整优化指令。

【当前使用的优化指令】
{optimize_template}

【所有尝试的完整评分详情】（共{len(all_attempt_logs)}次尝试，均未达标）
{full_history}

【低分维度】
{last_focus}

【元优化要求】
1. 分析为什么当前的优化策略反复尝试仍无法提升评分（是优化方向不对？约束不够？还是遗漏了关键因素？）
2. 生成一个新的优化指令，要求：
   - 保留当前指令中有效的部分
   - 修改或增加能解决反复低分问题的策略
   - 指令必须包含占位符 {{current_prompt}}、{{details}}、{{focus_text}}（用双花括号）
3. 输出完整的优化指令模板，不要省略
4. 不要输出任何解释说明，只输出新的优化指令模板本身"""

    try:
        logger.info(f"[optimize_method] 发送优化请求，内容长度={len(meta_instruction)} 字符，预计需要较长时间...")
        new_template = request_scoring_api(meta_instruction)
        logger.info(f"[optimize_method] 优化方法更新完成，原始长度={len(optimize_template)}, 新长度={len(new_template)}")
        # 保存优化方法到文件
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        method_file = os.path.join(LOG_DIR, f"optimized_method_{ts[:8]}.txt")
        with open(method_file, 'a', encoding='utf-8') as f:
            f.write(f"\n{'='*60}\n")
            f.write(f"===== 新的优化方法 ({ts}) =====\n")
            f.write(f"{'='*60}\n\n")
            f.write(new_template)
            f.write("\n")
        logger.info(f"[optimize_method] 新优化方法已保存到: {method_file}")
        return new_template
    except Exception as e:
        logger.error(f"优化方法更新失败: {e}")
        return optimize_template


def score_answer(question: str, ai_answer: str, reference_answer: str, scoring_prompt_template: str = None) -> dict:
    """使用AI对回答进行语义对比打分"""
    # 使用自定义提示词模板或默认模板
    template = scoring_prompt_template or DEFAULT_SCORING_PROMPT
    scoring_prompt = template.format(
        question=question,
        reference_answer=reference_answer,
        ai_answer=ai_answer
    )

    failed = {
        'success': False,
        'accuracy_score': 0, 'accuracy_reason': '评分失败',
        'citation_score': 0, 'citation_reason': '评分失败',
        'summary_score': 0, 'summary_reason': '评分失败',
        'total_score': 0
    }

    last_error = None
    for attempt in range(3):
        try:
            content = request_scoring_api(scoring_prompt)

            json_start = content.find('{')
            json_end = content.rfind('}') + 1
            if json_start == -1 or json_end <= json_start:
                last_error = f"评分返回无JSON: {content[:200]}"
                logger.error(f"第{attempt+1}次尝试 - {last_error}")
                continue

            scores = json.loads(content[json_start:json_end])
            score_keys = ['accuracy_score', 'citation_score', 'summary_score']
            result = {
                'success': True,
                **{k: scores.get(k, 0) for k in score_keys},
                **{k.replace('score', 'reason'): scores.get(k.replace('score', 'reason'), '') for k in score_keys},
                'total_score': sum(scores.get(k, 0) for k in score_keys)
            }
            logger.info(f"[score_answer] 评分完成 → 总分={result['total_score']}, 准确性={result['accuracy_score']}, 法条={result['citation_score']}, 总结={result['summary_score']}")
            return result
        except Exception as e:
            last_error = str(e)
            logger.error(f"第{attempt+1}次评分失败: {e}")
            if attempt < 2:
                import time
                time.sleep(2)

    # 3次都失败，把具体错误写入reason
    failed['accuracy_reason'] = f'评分失败({last_error})' if last_error else '评分失败'
    failed['citation_reason'] = failed['accuracy_reason']
    failed['summary_reason'] = failed['accuracy_reason']
    return failed


def read_questions_from_excel(filepath):
    """从Excel文件读取问题(B列)、现有AI回答(C列)和建议答案(E列)"""
    wb = load_workbook(filepath)
    ws = wb.active
    questions = []
    row = 2

    while True:
        cell_value = ws[f'B{row}'].value
        if cell_value is None or str(cell_value).strip() == '':
            break

        existing_answer = ws[f'C{row}'].value
        existing_answer = str(existing_answer).strip() if existing_answer else ''

        reference_answer = ws[f'E{row}'].value
        reference_answer = str(reference_answer).strip() if reference_answer else ''
        questions.append((row, str(cell_value).strip(), existing_answer, reference_answer))
        row += 1

    return questions


def save_results_to_excel(questions_with_answers, output_filepath):
    """保存结果到新Excel，含轮次信息，保留问题、建议答案、评分情况"""
    wb = Workbook()
    ws = wb.active
    ws.title = '评估结果'

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        ('轮次', 20),
        ('问题', 50),
        ('AI回答', 100),
        ('建议答案', 80),
        ('答案准确性(60分)', 12), ('答案准确性说明', 40),
        ('法条援引度(20分)', 12), ('法条援引度说明', 40),
        ('总结完整度(20分)', 12), ('总结完整度说明', 40),
        ('总分(100分)', 10)
    ]

    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, (round_label, question, answer, reference_answer, scores) in enumerate(questions_with_answers, 2):
        # 轮次
        cell = ws.cell(row=row_idx, column=1, value=round_label)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border

        # 问题
        cell = ws.cell(row=row_idx, column=2, value=question)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border

        # AI回答
        cell = ws.cell(row=row_idx, column=3, value=answer)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border

        # 建议答案
        cell = ws.cell(row=row_idx, column=4, value=reference_answer or '')
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border

        if scores:
            score_data = [
                scores.get('accuracy_score', 0),
                scores.get('accuracy_reason', ''),
                scores.get('citation_score', 0),
                scores.get('citation_reason', ''),
                scores.get('summary_score', 0),
                scores.get('summary_reason', ''),
                scores.get('total_score', 0)
            ]
            for i, value in enumerate(score_data):
                cell = ws.cell(row=row_idx, column=5 + i, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = thin_border

    wb.save(output_filepath)


def save_results_to_html(questions_with_answers, output_filepath):
    """保存结果到HTML文件，PwC风格，含轮次信息"""
    rows_html = ''
    for row_idx, (round_label, question, answer, reference_answer, scores) in enumerate(questions_with_answers, 1):
        bg = '#FFFFFF' if row_idx % 2 == 1 else '#F4F4F4'
        scores_html = ''
        if scores and scores.get('success'):
            score_items = [
                ('答案准确性', scores.get('accuracy_score', 0), scores.get('accuracy_reason', '')),
                ('法条援引度', scores.get('citation_score', 0), scores.get('citation_reason', '')),
                ('总结完整度', scores.get('summary_score', 0), scores.get('summary_reason', '')),
            ]
            items_html = ''
            for label, val, reason in score_items:
                color = '#D04A02' if val < 15 else '#2A9D8F'
                items_html += f'''<div style="display:inline-block;width:19%;text-align:center;padding:6px 0;">
                    <div style="font-size:11px;color:#8A8C8E;">{label}</div>
                    <div style="font-size:20px;font-weight:700;color:{color};">{val}/20</div>
                    <div style="font-size:11px;color:#8A8C8E;text-align:left;line-height:1.3;">{reason}</div>
                </div>'''
            total = scores.get('total_score', 0)
            total_color = '#D04A02' if total < 75 else '#2A9D8F'
            items_html += f'''<div style="display:inline-block;width:19%;text-align:center;padding:6px 0;background:#2D2D2D;border-radius:3px;color:white;">
                <div style="font-size:11px;color:rgba(255,255,255,0.7);">总分</div>
                <div style="font-size:20px;font-weight:700;color:{total_color};">{total}/100</div>
            </div>'''
            scores_html = f'<div style="margin-top:8px;padding:8px;background:white;border-radius:3px;border:1px solid #E8E8E8;">{items_html}</div>'
        elif scores:
            scores_html = f'<div style="margin-top:8px;padding:8px;background:#FEF2F2;border-radius:3px;color:#CB333B;font-size:13px;">评分失败</div>'

        safe_round = round_label.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        safe_q = question.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        safe_ans = answer.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        safe_ref = (reference_answer or '').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

        rows_html += f'''<tr style="background:{bg};vertical-align:top;">
            <td style="padding:10px;border:1px solid #E8E8E8;width:10%;">{safe_round}</td>
            <td style="padding:10px;border:1px solid #E8E8E8;width:10%;">{safe_q}</td>
            <td style="padding:10px;border:1px solid #E8E8E8;width:28%;">{safe_ans}</td>
            <td style="padding:10px;border:1px solid #E8E8E8;width:16%;">{safe_ref}</td>
            <td style="padding:10px;border:1px solid #E8E8E8;width:36%;">{scores_html}</td>
        </tr>'''

    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>AI回答评估结果</title>
<style>
body {{ font-family: Helvetica, Arial, 'PingFang SC', 'Microsoft YaHei', sans-serif; background:#F4F4F4; padding:20px; color:#2D2D2D; }}
h1 {{ font-size:20px; font-weight:600; border-bottom:2px solid #D04A02; padding-bottom:10px; }}
</style>
</head>
<body>
<h1>AI回答评估结果</h1>
<table style="width:100%;border-collapse:collapse;font-size:13px;">
<thead><tr style="background:#2D2D2D;color:white;">
<th style="padding:10px;border:1px solid #2D2D2D;width:10%;text-align:left;">轮次</th>
<th style="padding:10px;border:1px solid #2D2D2D;width:10%;text-align:left;">问题</th>
<th style="padding:10px;border:1px solid #2D2D2D;width:28%;text-align:left;">AI回答</th>
<th style="padding:10px;border:1px solid #2D2D2D;width:16%;text-align:left;">建议答案</th>
<th style="padding:10px;border:1px solid #2D2D2D;width:36%;text-align:left;">评分情况</th>
</tr></thead>
<tbody>{rows_html}</tbody>
</table>
</body></html>'''

    with open(output_filepath, 'w', encoding='utf-8') as f:
        f.write(html)


@app.route('/')
def index():
    return render_template('index.html')


# 提示词持久化文件路径
SCORING_PROMPT_FILE = os.path.join(ROOT_DIR, 'scoring_prompt.txt')


def load_saved_prompt():
    """加载已保存的评分提示词，如果没有则返回默认模板"""
    if os.path.exists(SCORING_PROMPT_FILE):
        try:
            with open(SCORING_PROMPT_FILE, 'r', encoding='utf-8') as f:
                saved = f.read().strip()
                if saved:
                    return saved
        except Exception as e:
            logger.warning(f"读取保存的提示词失败: {e}")
    return DEFAULT_SCORING_PROMPT


def save_prompt(prompt_text):
    """保存评分提示词到文件"""
    with open(SCORING_PROMPT_FILE, 'w', encoding='utf-8') as f:
        f.write(prompt_text)


@app.route('/default_prompt', methods=['GET'])
def get_default_prompt():
    """返回默认评分提示词模板"""
    return jsonify({'prompt': DEFAULT_SCORING_PROMPT})


@app.route('/saved_prompt', methods=['GET'])
def get_saved_prompt():
    """返回已保存的评分提示词（优先返回用户自定义，没有则返回默认）"""
    return jsonify({'prompt': load_saved_prompt()})


@app.route('/save_prompt', methods=['POST'])
def save_prompt_api():
    """保存用户编辑的评分提示词"""
    data = request.json
    prompt_text = data.get('prompt', '').strip()
    if not prompt_text:
        return jsonify({'error': '提示词不能为空'}), 400
    try:
        save_prompt(prompt_text)
        return jsonify({'success': True, 'message': '提示词已保存'})
    except Exception as e:
        return jsonify({'error': f'保存失败: {str(e)}'}), 500


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': '没有上传文件'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400

    if not (file and allowed_file(file.filename)):
        return jsonify({'error': '不支持的文件格式'}), 400

    # 保留中文文件名，只去掉不安全字符
    safe_name = re.sub(r'[\\/*?:"<>|]', '_', file.filename)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{timestamp}_{safe_name}"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    try:
        questions = read_questions_from_excel(filepath)
    except Exception as e:
        return jsonify({'error': f'读取Excel文件失败: {str(e)}'}), 400

    return jsonify({
        'success': True,
        'filename': filename,
        'question_count': len(questions),
        'questions': [{'row': r, 'question': q, 'has_existing_answer': bool(ea), 'has_reference': bool(ref)} for r, q, ea, ref in questions]
    })


def process_single_question(row_num, question, reference_answer, enable_scoring, scoring_prompt_template=None, system_prompt="", existing_answer=""):
    """处理单个问题：获取AI回答 + 评分。如果回答为空则自动重试"""
    try:
        if existing_answer:
            answer = existing_answer
        else:
            max_retries = 2
            for retry in range(max_retries + 1):
                answer = chat_with_confirmation(question, system_prompt=system_prompt)
                if answer.strip():
                    break
                if retry < max_retries:
                    logger.info(f"[process] row={row_num} 回答为空，重试第{retry+1}次...")

        scores = None
        if enable_scoring and reference_answer:
            logger.info(f"[process] row={row_num} 开始评分...")
            scores = score_answer(question, answer, reference_answer, scoring_prompt_template)
            if scores:
                logger.info(f"[process] row={row_num} 评分完成 → 总分={scores.get('total_score', '?')}")
            else:
                logger.warning(f"[process] row={row_num} 评分返回None")

        return {
            'row': row_num,
            'question': question,
            'answer': answer,
            'reference_answer': reference_answer,
            'scores': scores,
            'success': True
        }
    except Exception as e:
        logger.error(f"问题处理失败 (行{row_num}): {e}")
        return {
            'row': row_num,
            'question': question,
            'answer': f'处理失败: {str(e)}',
            'reference_answer': reference_answer,
            'scores': None,
            'success': False
        }


def _log_attempt_summary(attempt, results, suffix="", context=""):
    """记录单次尝试的日志：级次信息 + 每题评分 + 整体统计，返回平均总分"""
    header = f"===== 第{attempt}次尝试{suffix}"
    if context:
        header += f" [{context}]"
    logger.info(header)

    scored = [r for r in results if r.get('scores') and r['scores'].get('success')]
    for r in results:
        scores = r.get('scores') or {}
        if scores.get('success'):
            logger.info(
                f"  题目(row={r['row']}): 总分={scores['total_score']}, "
                f"准确性={scores['accuracy_score']}/60, "
                f"法条援引={scores['citation_score']}/20, "
                f"总结完整度={scores['summary_score']}/20"
            )
        else:
            logger.info(f"  题目(row={r['row']}): 评分失败")

    if scored:
        n = len(scored)
        avg_total = sum(r['scores']['total_score'] for r in scored) / n
        avg_acc = sum(r['scores']['accuracy_score'] for r in scored) / n
        avg_cit = sum(r['scores']['citation_score'] for r in scored) / n
        avg_sum = sum(r['scores']['summary_score'] for r in scored) / n
        logger.info(
            f"整体统计: 平均总分={avg_total:.1f}, "
            f"平均准确性={avg_acc:.1f}, "
            f"平均法条援引={avg_cit:.1f}, "
            f"平均总结完整度={avg_sum:.1f}"
        )
        return avg_total
    return 0


@app.route('/process', methods=['POST'])
def process_questions():
    """SSE流式处理接口，支持两层自动优化（内层优化提示词，外层优化优化方法）"""
    data = request.json
    filename = data.get('filename')
    enable_scoring = data.get('enable_scoring', False)
    scoring_prompt_template = data.get('scoring_prompt', None)
    thread_count = data.get('thread_count', 2)
    thread_count = max(1, min(8, int(thread_count)))
    max_attempts = max(1, min(20, int(data.get('max_attempts', get_config('max_attempts', 10)))))
    max_optimize_rounds = max(1, min(5, int(data.get('max_optimize_rounds', get_config('max_optimize_rounds', 3)))))
    score_threshold = get_config('score_threshold', 80)

    def sse_error(msg):
        def gen():
            yield f"data: {json.dumps({'type': 'error', 'message': msg}, ensure_ascii=False)}\n\n"
        return app.response_class(gen(), mimetype='text/event-stream',
                                  headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})

    if not filename:
        return sse_error('缺少文件名')

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(filepath):
        return sse_error('文件不存在')

    questions = read_questions_from_excel(filepath)
    total = len(questions)
    system_prompt = load_system_prompt()
    optimize_template = None  # 使用内层默认的优化指令
    logger.info(f"========== 开始处理 {total}题 | {thread_count}线程 | 评分{'开' if enable_scoring else '关'} | 内层{max_attempts}次 | 外层{max_optimize_rounds}轮 | 阈值{score_threshold} ==========")

    def generate():
        nonlocal system_prompt, optimize_template

        global_attempt = 0
        all_round_logs = []

        for optimize_round in range(1, max_optimize_rounds + 1):
            logger.info(f"{'#'*60}")
            logger.info(f"### 外层第{optimize_round}/{max_optimize_rounds}轮（优化方法{'：默认' if optimize_round == 1 else '：已更新'}）###")
            logger.info(f"{'#'*60}")

            all_attempt_logs = []
            results = []
            eval_rounds = max(1, int(get_config('eval_rounds', 3)))
            current_prompt_avg = 0.0  # 当前提示词版本的多次评估平均分

            for inner_attempt in range(1, max_attempts + 1):
                global_attempt += 1
                try:
                    if inner_attempt == 1:
                        logger.info(f"===== 内层第{inner_attempt}次尝试（每个版本评估{eval_rounds}次取平均）=====")
                    else:
                        logger.info(f"===== 内层第{inner_attempt}次尝试（使用优化后提示词，评估{eval_rounds}次取平均）=====")

                    # 同一提示词跑 eval_rounds 次，取平均分
                    round_scores = []
                    last_results = []

                    for eval_i in range(1, eval_rounds + 1):
                        if eval_rounds > 1:
                            logger.info(f"--- 第{eval_i}/{eval_rounds}轮评估 ---")

                        result_queue = queue.Queue()

                        def worker(row_num, question, _existing_answer, reference_answer):
                            result = process_single_question(
                                row_num, question, reference_answer,
                                enable_scoring, scoring_prompt_template, system_prompt
                            )
                            result_queue.put(result)

                        with ThreadPoolExecutor(max_workers=thread_count) as executor:
                            futures = []
                            for row_num, question, existing_answer, reference_answer in questions:
                                futures.append(executor.submit(worker, row_num, question, existing_answer, reference_answer))

                            completed = 0
                            eval_results = []
                            while completed < total:
                                result = result_queue.get()
                                eval_results.append(result)
                                completed += 1

                                event = {
                                    'type': 'progress',
                                    'current': completed,
                                    'total': total,
                                    'percentage': int(completed / total * 100),
                                    'attempt': global_attempt,
                                    'max_attempts': max_attempts,
                                    'optimize_round': optimize_round,
                                    'eval_i': eval_i,
                                    'eval_rounds': eval_rounds,
                                    'result': result
                                }
                                yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"

                        eval_results.sort(key=lambda r: r['row'])
                        eval_avg = _log_attempt_summary(
                            global_attempt,
                            eval_results,
                            suffix=f" [评估{eval_i}/{eval_rounds}]" if eval_rounds > 1 else "",
                            context=f"外层{optimize_round}轮-内层{inner_attempt}次"
                        )
                        round_scores.append(eval_avg)
                        last_results = eval_results

                    # 取多次评估的平均分
                    current_prompt_avg = sum(round_scores) / len(round_scores)
                    if eval_rounds > 1:
                        logger.info(f">>> 当前提示词{eval_rounds}轮评估平均分: {current_prompt_avg:.1f}（{', '.join(f'{s:.1f}' for s in round_scores)}）")

                    all_attempt_logs.append({
                        'attempt': global_attempt,
                        'inner_attempt': inner_attempt,
                        'system_prompt': system_prompt,
                        'avg_total': current_prompt_avg,
                        'eval_scores': round_scores,
                        'results': last_results
                    })

                    # 检查是否需要优化：平均分不达标 或 存在答案准确性低分
                    min_accuracy_score = get_config('min_accuracy_score', 40)
                    low_accuracy_questions = [
                        r for r in last_results
                        if r.get('scores') and r['scores'].get('success')
                        and r['scores'].get('accuracy_score', 0) < min_accuracy_score
                    ]
                    has_low_accuracy = len(low_accuracy_questions) > 0
                    low_accuracy_info = ""
                    if has_low_accuracy:
                        low_accuracy_info = f"，其中{len(low_accuracy_questions)}题准确性低于{min_accuracy_score}分（" + \
                            "、".join(f"第{r['row']}题={r['scores']['accuracy_score']}分" for r in low_accuracy_questions) + "）"

                    # 平均分达标且无准确性低分，直接结束
                    if current_prompt_avg >= score_threshold and not has_low_accuracy:
                        logger.info(f">>> 平均总分 {current_prompt_avg:.1f} >= {score_threshold}，且无准确性低于{min_accuracy_score}分，达标！")
                        break

                    need_retry = (
                        enable_scoring
                        and inner_attempt < max_attempts
                        and (current_prompt_avg < score_threshold or has_low_accuracy)
                        and current_prompt_avg >= 0
                        and len([r for r in last_results if r.get('scores') and r['scores'].get('success')]) > 0
                    )

                    if need_retry:
                        reason = f"平均总分 {current_prompt_avg:.1f} < {score_threshold}"
                        if has_low_accuracy:
                            reason += f" + {len(low_accuracy_questions)}题准确性低于{min_accuracy_score}分"
                        logger.info(f">>> {reason}{low_accuracy_info}，触发第{inner_attempt}次提示词优化...")
                        yield f"data: {json.dumps({'type': 'optimizing', 'attempt': global_attempt, 'avg_score': round(current_prompt_avg, 1)}, ensure_ascii=False)}\n\n"

                        old_prompt = system_prompt
                        ctx = f"外层第{optimize_round}轮 | 内层第{inner_attempt}次尝试 | 优化前平均分={current_prompt_avg:.1f} | 原始长度={len(system_prompt)}"
                        system_prompt = optimize_prompt(system_prompt, last_results, global_attempt, optimize_template, ctx)

                        if system_prompt != old_prompt:
                            logger.info(f">>> 提示词已优化，将用新提示词重新获取AI回答并评分")
                        else:
                            logger.info(">>> 提示词未变更，终止内层循环")
                            break
                    else:
                        break

                except Exception as e:
                    logger.error(f"第{global_attempt}次尝试异常: {e}")
                    yield f"data: {json.dumps({'type': 'error', 'message': f'第{global_attempt}次尝试异常: {str(e)}'}, ensure_ascii=False)}\n\n"
                    if not results:
                        return

            # 内层循环结束，记录本轮
            round_best_avg = max((e['avg_total'] for e in all_attempt_logs), default=0)
            all_round_logs.append({
                'round': optimize_round,
                'best_avg': round_best_avg,
                'attempts': len(all_attempt_logs),
                'all_attempt_logs': all_attempt_logs
            })

            # 已达标则不进入外层循环
            if round_best_avg >= score_threshold:
                break

            # 外层还有轮次，优化优化方法（带异常保护）
            if optimize_round < max_optimize_rounds and round_best_avg > 0:
                try:
                    logger.info(f"{'#'*60}")
                    logger.info(f"### 内层{len(all_attempt_logs)}次尝试后仍 < {score_threshold}，开始优化优化方法... ###")
                    logger.info(f"{'#'*60}")
                    old_template = optimize_template or ""
                    optimize_template = optimize_optimization_method(
                        optimize_template or "默认优化指令", all_attempt_logs
                    )
                    if optimize_template == old_template:
                        logger.info(">>> 优化方法未变更，终止外层循环")
                        break
                    # 重置 system_prompt 为最高评分版本
                    best = max(all_attempt_logs, key=lambda x: x['avg_total'])
                    system_prompt = best['system_prompt']
                    logger.info(f">>> 优化方法已更新，重置系统提示词为最高评分版（平均{best['avg_total']:.1f}），开始下一轮")
                except Exception as e:
                    logger.error(f"外层优化方法异常: {e}，终止外层循环")
                    break

        # 保存最终结果（所有轮次，按轮次区分）
        try:
            all_results = []
            for rl in all_round_logs:
                round_label = f"外层{rl['round']}轮"
                for att_log in rl['all_attempt_logs']:
                    att_label = f"{round_label}-内层{att_log['inner_attempt']}次(平均{att_log['avg_total']:.1f})"
                    for r in sorted(att_log['results'], key=lambda x: x['row']):
                        all_results.append((att_label, r['question'], r['answer'], r.get('reference_answer', ''), r['scores']))

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename_xlsx = f"AI评估结果_{timestamp}.xlsx"
            output_filename_html = f"AI评估结果_{timestamp}.html"
            output_xlsx = os.path.join(app.config['OUTPUT_FOLDER'], output_filename_xlsx)
            output_html = os.path.join(app.config['OUTPUT_FOLDER'], output_filename_html)
            logger.info(f"保存结果 → Excel={output_filename_xlsx}, HTML={output_filename_html}（共{len(all_results)}行）")
            save_results_to_excel(all_results, output_xlsx)
            save_results_to_html(all_results, output_html)
        except Exception as e:
            logger.error(f"保存结果异常: {e}")
            output_filename_xlsx = ""
            output_filename_html = ""

        # 最终汇总
        try:
            logger.info("=" * 60)
            # 汇总所有尝试
            all_attempts = []
            for rl in all_round_logs:
                all_attempts.extend(rl['all_attempt_logs'])
            if len(all_attempts) > 1:
                first_avg = all_attempts[0]['avg_total']
                best = max(all_attempts, key=lambda x: x['avg_total'])
                logger.info(f"各次尝试平均总分: " + " → ".join(f"第{e['attempt']}次={e['avg_total']:.1f}" for e in all_attempts))
                logger.info(f"评分提升: {first_avg:.1f} → {best['avg_total']:.1f} (提高了 {best['avg_total'] - first_avg:.1f} 分)")
                inner_att = best.get('inner_attempt', '?')
                # 找到该attempt所属的外层轮次
                best_round = '?'
                for rl in all_round_logs:
                    for a in rl['all_attempt_logs']:
                        if a['attempt'] == best['attempt']:
                            best_round = rl['round']
                            break
                logger.info("")
                logger.info("=" * 60)
                logger.info("===== 保存最优提示词开始 =====")
                logger.info(f"外层第{best_round}轮-内层第{inner_att}次, 第{best['attempt']}次全局, 平均总分={best['avg_total']:.1f}")
                logger.info("=" * 60)
                logger.info(best['system_prompt'])
                logger.info("=" * 60)
                logger.info("===== 保存最优提示词结束 =====")
                logger.info("=" * 60)
            logger.info(f"========== 全部完成，共{global_attempt}次尝试（{len(all_round_logs)}轮优化方法），结果已保存 ==========")
        except Exception as e:
            logger.error(f"汇总日志异常: {e}")

        yield f"data: {json.dumps({'type': 'complete', 'output_filename': output_filename_xlsx, 'output_filename_html': output_filename_html, 'attempts': global_attempt}, ensure_ascii=False)}\n\n"

    return app.response_class(generate(), mimetype='text/event-stream',
                              headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


@app.route('/quick_score', methods=['POST'])
def quick_score():
    """快速评分接口 - 手动输入问题和答案直接评分"""
    data = request.json
    question = (data.get('question') or '').strip()
    ai_answer = (data.get('ai_answer') or '').strip()
    reference_answer = (data.get('reference_answer') or '').strip()
    scoring_prompt_template = data.get('scoring_prompt') or None

    if not question or not ai_answer or not reference_answer:
        return jsonify({'success': False, 'error': '问题、AI回答、建议回答均为必填'}), 400

    try:
        scores = score_answer(question, ai_answer, reference_answer, scoring_prompt_template)
        return jsonify({'success': True, 'scores': scores})
    except Exception as e:
        logger.error(f"快速评分失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/evaluate', methods=['POST'])
def api_evaluate():
    """
    外部API接口 - 批量评估问题
    入参JSON格式：
    {
        "items": [
            {"question": "问题1", "reference_answer": "建议答案1"},
            {"question": "问题2", "reference_answer": "建议答案2"}
        ],
        "scoring_prompt": "可选，自定义评分提示词模板",
        "thread_count": 4
    }
    返回JSON格式：
    {
        "success": true,
        "count": 2,
        "results": [
            {
                "question": "问题1",
                "answer": "AI回答内容",
                "reference_answer": "建议答案1",
                "scores": { ... }
            }
        ]
    }
    """
    data = request.json
    if not data or 'items' not in data:
        return jsonify({'success': False, 'error': '请求体需包含 items 数组'}), 400

    items = data.get('items', [])
    if not items or not isinstance(items, list):
        return jsonify({'success': False, 'error': 'items 不能为空'}), 400

    # 校验每条数据
    for i, item in enumerate(items):
        if not item.get('question'):
            return jsonify({'success': False, 'error': f'第{i+1}条数据缺少 question 字段'}), 400

    scoring_prompt_template = data.get('scoring_prompt') or load_saved_prompt()
    thread_count = max(1, min(8, int(data.get('thread_count', 4))))
    max_attempts = max(1, min(20, int(data.get('max_attempts', 20))))
    system_prompt = load_system_prompt()

    def process_one(idx, item):
        question = item['question']
        reference_answer = item.get('reference_answer', '')
        try:
            answer = chat_with_confirmation(question, system_prompt=system_prompt)
            scores = None
            if reference_answer:
                scores = score_answer(question, answer, reference_answer, scoring_prompt_template)
            return {
                'index': idx,
                'question': question,
                'answer': answer,
                'reference_answer': reference_answer,
                'scores': scores,
                'success': True
            }
        except Exception as e:
            logger.error(f"API评估失败 (第{idx+1}条): {e}")
            return {
                'index': idx,
                'question': question,
                'answer': f'处理失败: {str(e)}',
                'reference_answer': reference_answer,
                'scores': None,
                'success': False
            }

    all_attempt_logs = []
    logger.info(f"[api_evaluate] 开始: {len(items)}题, max_attempts={max_attempts}, threads={thread_count}")

    for attempt in range(1, max_attempts + 1):
        logger.info(f"[api_evaluate] ===== 第{attempt}/{max_attempts}次尝试 =====")
        results = [None] * len(items)
        with ThreadPoolExecutor(max_workers=thread_count) as executor:
            futures = {executor.submit(process_one, i, item): i for i, item in enumerate(items)}
            for future in futures:
                result = future.result()
                results[result['index']] = result

        avg_total = _log_attempt_summary(attempt, results)
        all_attempt_logs.append({
            'attempt': attempt,
            'system_prompt': system_prompt,
            'avg_total': avg_total
        })

        need_retry = attempt < max_attempts and avg_total < 70 and len([r for r in results if r.get('scores', {}).get('success')]) > 0
        if need_retry:
            logger.info(f"api_evaluate: 平均总分 {avg_total:.1f} < 70，优化提示词...")
            old_prompt = system_prompt
            ctx = f"api_evaluate | 第{attempt}次尝试 | 优化前平均分={avg_total:.1f} | 原始长度={len(system_prompt)}"
            system_prompt = optimize_prompt(system_prompt, results, attempt, context_info=ctx)
            if system_prompt == old_prompt:
                logger.info("提示词未变更，终止重试")
                break
        else:
            break

    # 记录各次尝试对比
    if len(all_attempt_logs) > 1:
        logger.info("===== api_evaluate 各次尝试评分对比 =====")
        for log_entry in all_attempt_logs:
            logger.info(
                f"第{log_entry['attempt']}次: 平均总分={log_entry['avg_total']:.1f}, "
                f"提示词前200字={log_entry['system_prompt'][:200]}"
            )

        best = max(all_attempt_logs, key=lambda x: x['avg_total'])
        logger.info("")
        logger.info("=" * 60)
        logger.info("===== 保存最优提示词开始 =====")
        logger.info(f"第{best['attempt']}次, 平均总分={best['avg_total']:.1f}")
        logger.info("=" * 60)
        logger.info(best['system_prompt'])
        logger.info("=" * 60)
        logger.info("===== 保存最优提示词结束 =====")
        logger.info("=" * 60)

    logger.info(f"[api_evaluate] 全部完成，共尝试{attempt}次")

    return jsonify({
        'success': True,
        'count': len(results),
        'results': results,
        'attempts': attempt
    })


@app.route('/download/<path:filename>')
def download_file(filename):
    filepath = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return jsonify({'error': '文件不存在'}), 404


if __name__ == '__main__':
    import sys
    if ROOT_DIR not in sys.path:
        sys.path.insert(0, ROOT_DIR)
    from app.model_scoring import model_scoring_bp
    app.register_blueprint(model_scoring_bp)
    app.run(debug=True, port=5001, host='0.0.0.0')
