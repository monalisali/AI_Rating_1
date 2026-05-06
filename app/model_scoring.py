#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
模型打分模块 - 独立评分，不做提示词优化
每道题使用Excel中自带的提示词，调用5007 API获取AI回答后评分
"""

import os
import json
import queue
import logging
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor

from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side

# 复用 app.py 中的函数
from app.app import (
    chat_with_confirmation, score_answer, get_config, logger,
    ROOT_DIR, allowed_file, DEFAULT_SCORING_PROMPT
)

model_scoring_bp = Blueprint('model_scoring', __name__)

OUTPUT_FOLDER = os.path.join(ROOT_DIR, 'outputs')
UPLOAD_FOLDER = os.path.join(ROOT_DIR, 'uploads')
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


@model_scoring_bp.route('/model_scoring_upload', methods=['POST'])
def model_scoring_upload():
    """上传Excel文件，解析问题、提示词、建议回答"""
    file = request.files.get('file')
    if not file or not allowed_file(file.filename):
        return jsonify({'error': '请上传 .xlsx 或 .xls 文件'}), 400

    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        questions = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            question = str(row[0]).strip() if len(row) > 0 and row[0] else ''
            prompt = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            ref_answer = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            if not question or question == 'None':
                continue
            questions.append({
                'row': row_idx,
                'question': question,
                'prompt': prompt,
                'reference_answer': ref_answer,
                'has_prompt': bool(prompt),
                'has_reference': bool(ref_answer)
            })
        wb.close()
    except Exception as e:
        return jsonify({'error': f'解析Excel失败: {e}'}), 400

    if not questions:
        return jsonify({'error': '未找到有效数据，请确保A列有问题'}), 400

    return jsonify({
        'filename': file.filename,
        'question_count': len(questions),
        'questions': questions
    })


def _process_one_question(row_num, question, prompt, reference_answer, scoring_prompt_template):
    """处理单个问题：用指定提示词获取AI回答 + 评分"""
    try:
        logger.info(f"[model_scoring] row={row_num} 开始调用5007 API获取AI回答...")
        answer = chat_with_confirmation(question, system_prompt=prompt)
        logger.info(f"[model_scoring] row={row_num} AI回答完成（长度={len(answer)}）")

        scores = None
        if reference_answer:
            logger.info(f"[model_scoring] row={row_num} 开始评分...")
            scores = score_answer(question, answer, reference_answer, scoring_prompt_template)
            if scores:
                logger.info(f"[model_scoring] row={row_num} 评分完成 → 总分={scores.get('total_score', '?')}, 准确性={scores.get('accuracy_score', '?')}, 法条={scores.get('citation_score', '?')}, 总结={scores.get('summary_score', '?')}")
        else:
            logger.info(f"[model_scoring] row={row_num} 无参考答案，跳过评分")

        return {
            'row': row_num,
            'question': question,
            'answer': answer,
            'reference_answer': reference_answer,
            'prompt_used': prompt,
            'scores': scores,
            'success': True
        }
    except Exception as e:
        logger.error(f"[model_scoring] 问题处理失败 (行{row_num}): {e}")
        return {
            'row': row_num,
            'question': question,
            'answer': f'处理失败: {str(e)}',
            'reference_answer': reference_answer,
            'prompt_used': prompt,
            'scores': None,
            'success': False
        }


@model_scoring_bp.route('/model_scoring_process', methods=['POST'])
def model_scoring_process():
    """处理所有问题，SSE流式返回进度"""
    data = request.get_json()
    filename = data.get('filename', '')
    rounds = max(1, int(data.get('rounds', 1)))
    scoring_prompt_template = data.get('scoring_prompt', '') or DEFAULT_SCORING_PROMPT
    thread_count = max(1, int(data.get('thread_count', 2)))

    filepath = os.path.join(UPLOAD_FOLDER, filename)
    if not os.path.exists(filepath):
        return jsonify({'error': '文件不存在，请重新上传'}), 400

    # 解析Excel
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        questions = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            question = str(row[0]).strip() if len(row) > 0 and row[0] else ''
            prompt = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            ref_answer = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            if not question or question == 'None':
                continue
            questions.append((row_idx, question, prompt, ref_answer))
        wb.close()
    except Exception as e:
        return jsonify({'error': f'解析Excel失败: {e}'}), 400

    total = len(questions)
    if total == 0:
        return jsonify({'error': '未找到有效问题'}), 400

    model_name = get_config('model', 'glm-5.1')
    logger.info(f"[model_scoring] 开始处理 {total}题 × {rounds}轮 | {thread_count}线程 | 模型={model_name}")

    def generate():
        import threading
        all_round_results = []  # 每轮的结果

        for round_i in range(1, rounds + 1):
            if rounds > 1:
                logger.info(f"[model_scoring] ===== 第{round_i}/{rounds}轮 =====")
            result_queue = queue.Queue()
            timeout_per_question = 1200  # 单题超时20分钟

            def worker(row_num, question, prompt, ref_answer):
                """带超时保护的单题处理：超时返回失败结果，不阻塞其他题目"""
                result_holder = [None]

                def _run():
                    try:
                        result_holder[0] = _process_one_question(row_num, question, prompt, ref_answer, scoring_prompt_template)
                    except Exception as e:
                        result_holder[0] = {
                            'row': row_num, 'question': question,
                            'answer': f'处理失败: {str(e)}',
                            'reference_answer': ref_answer,
                            'prompt_used': prompt,
                            'scores': None, 'success': False
                        }

                t = threading.Thread(target=_run, daemon=True)
                t.start()
                t.join(timeout=timeout_per_question)

                if t.is_alive():
                    logger.warning(f"[model_scoring] row={row_num} 第{round_i}轮超时（>{timeout_per_question}秒），跳过")
                    result_queue.put({
                        'row': row_num, 'question': question,
                        'answer': f'超时（>{timeout_per_question}秒），已跳过',
                        'reference_answer': ref_answer,
                        'prompt_used': prompt,
                        'scores': None, 'success': False
                    })
                else:
                    result_queue.put(result_holder[0])

            with ThreadPoolExecutor(max_workers=thread_count) as executor:
                futures = []
                for row_num, question, prompt, ref_answer in questions:
                    futures.append(executor.submit(worker, row_num, question, prompt, ref_answer))

                completed = 0
                round_results = []
                while completed < total:
                    result = result_queue.get()
                    round_results.append(result)
                    completed += 1

                    event = {
                        'type': 'progress',
                        'current': completed,
                        'total': total,
                        'percentage': int(completed / total * 100),
                        'round': round_i,
                        'rounds': rounds,
                        'result': result
                    }
                    yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"

            round_results.sort(key=lambda r: r['row'])
            all_round_results.append({'round': round_i, 'results': round_results})

            # 本轮统计
            scored = [r for r in round_results if r.get('scores') and r['scores'].get('success')]
            if scored:
                avg_total = sum(r['scores']['total_score'] for r in scored) / len(scored)
                logger.info(f"[model_scoring] 第{round_i}轮完成 → 已评分{len(scored)}/{total}题，平均总分={avg_total:.1f}")
            else:
                logger.info(f"[model_scoring] 第{round_i}轮完成 → 无有效评分")

        # 保存Excel到 outputs/模型打分结果/ 子文件夹
        model_output_dir = os.path.join(OUTPUT_FOLDER, '模型打分结果')
        os.makedirs(model_output_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"模型打分结果_{timestamp}.xlsx"
        output_path = os.path.join(model_output_dir, output_filename)
        _save_to_excel(all_round_results, output_path, model_name)
        logger.info(f"[model_scoring] Excel已保存: {output_path}")
        logger.info(f"[model_scoring] ===== 全部完成 =====")

        # 完成事件（下载路径仍用 /download/ 路由，文件放在子目录下）
        yield f"data: {json.dumps({'type': 'complete', 'output_filename': f'模型打分结果/{output_filename}'}, ensure_ascii=False)}\n\n"

    return __import__('flask').Response(generate(), mimetype='text/event-stream')


def _save_to_excel(all_round_results, output_path, model_name):
    """保存模型打分结果到Excel，每轮一个区域，标注模型名称"""
    wb = Workbook()
    ws = wb.active
    ws.title = '模型打分结果'

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        ('轮次', 12),
        ('模型', 15),
        ('问题', 50),
        ('AI回答', 100),
        ('建议答案', 80),
        ('提示词', 40),
        ('答案准确性(60分)', 12), ('答案准确性说明', 40),
        ('法条援引度(20分)', 12), ('法条援引度说明', 40),
        ('总结完整度(20分)', 12), ('总结完整度说明', 40),
        ('总分(100分)', 10)
    ]

    # 表头
    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        ws.column_dimensions[cell.column_letter].width = width

    row_idx = 2
    for round_data in all_round_results:
        round_label = f"第{round_data['round']}轮"
        for r in round_data['results']:
            # 轮次
            cell = ws.cell(row=row_idx, column=1, value=round_label)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border

            # 模型
            cell = ws.cell(row=row_idx, column=2, value=model_name)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border

            # 问题
            cell = ws.cell(row=row_idx, column=3, value=r['question'])
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border

            # AI回答
            cell = ws.cell(row=row_idx, column=4, value=r['answer'])
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border

            # 建议答案
            cell = ws.cell(row=row_idx, column=5, value=r.get('reference_answer', ''))
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border

            # 提示词(前200字)
            cell = ws.cell(row=row_idx, column=6, value=r.get('prompt_used', ''))
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border

            # 评分
            scores = r.get('scores')
            if scores:
                score_data = [
                    scores.get('accuracy_score', 0),
                    scores.get('accuracy_reason', ''),
                    scores.get('citation_score', 0),
                    scores.get('citation_reason', ''),
                    scores.get('summary_score', 0),
                    scores.get('summary_reason', ''),
                    scores.get('total_score', 0)
                ]
                for i, value in enumerate(score_data):
                    cell = ws.cell(row=row_idx, column=7 + i, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = thin_border

            row_idx += 1

    wb.save(output_path)
