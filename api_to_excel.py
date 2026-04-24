#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
API请求脚本 - 获取AI回答并保存到Excel表格
支持多轮对话确认
支持从Excel文件读取问题（从B2开始按行读取）
"""

import json
import ssl
import sys
import urllib.request
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


def request_api(message: str, session_id: str = "") -> tuple:
    """
    请求API接口并返回content值和session_id

    Args:
        message: 用户问题
        session_id: 会话ID（可选）

    Returns:
        (api_response, session_id) 元组
    """
    url = 'https://ai.tech.tax.asia.pwcinternal.com:5007/api/chat-stream'
    data = json.dumps({
        'message': message,
        'session_id': session_id
    }).encode('utf-8')

    req = urllib.request.Request(
        url,
        data=data,
        headers={
            'Content-Type': 'application/json',
            'Accept': 'text/event-stream'
        }
    )

    # 配置SSL（跳过证书验证）
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    # 发送请求
    with urllib.request.urlopen(req, context=ctx, timeout=300) as response:
        # 从响应头获取session_id
        returned_session_id = response.headers.get('X-Session-Id', '')
        result = response.read().decode('utf-8')

    return result, returned_session_id


def parse_response(api_response: str) -> dict:
    """
    解析API响应

    Args:
        api_response: API返回的原始响应

    Returns:
        包含contents列表和是否结束的字典
    """
    lines = [l.strip() for l in api_response.strip().split('\n') if l.strip()]

    contents = []
    for line in lines:
        if line.startswith('data:'):
            json_str = line[5:].strip()
            try:
                data = json.loads(json_str)
                if data.get('type') == 'content':
                    contents.append(data.get('content', ''))
            except json.JSONDecodeError:
                pass

    return {
        'contents': contents,
        'full_content': ''.join(contents)
    }


def is_confirmation_question(content: str) -> bool:
    """
    判断内容是否是确认问句

    Args:
        content: AI返回的内容

    Returns:
        是否需要确认
    """
    confirmation_keywords = [
        '请问以上关键词是否需要调整或补充',
        '确认后我将开始',
        '请问这样理解是否正确',
        '您是否有需要补充或调整',
        '是否需要调整',
        '请确认',
        '确认后',
        '以上关键词是否准确',
        '是否需要添加',
        '请问以上',
        '确认后开始',
        '是否准确',
        '请告知',
        '是否继续',
        '我将开始',
        '以上内容是否',
    ]

    for keyword in confirmation_keywords:
        if keyword in content:
            return True
    return False


def remove_references(content: str) -> str:
    """
    移除引用文件部分的内容

    Args:
        content: 原始内容

    Returns:
        移除引用后的内容
    """
    result = content

    # 按优先级处理的引用部分标记（从后往前找第一个出现的）
    reference_markers = [
        # Markdown标题格式的引用部分
        '\n\n## 引用文件',
        '\n\n## 引用法规列表',
        '\n\n## 引用法规',
        '\n\n## 参考资料',
        '\n\n## 参考文献',
        '\n\n## 参考文件',
        '\n\n## 来源',
        '\n\n## 相关法规',
        # 加粗格式的引用部分
        '\n\n**参考资料**',
        '\n\n**参考文件**',
        '\n\n**引用文件**',
        '\n\n**引用文件及链接**',
        '\n\n**来源**',
        # 分隔符 + 引用
        '\n\n---\n\n**参考',
        '\n\n---\n\n**温馨提示',
        '\n\n---\n**参考',
        '\n\n---\n**温馨提示',
        '\n---\n\n**参考',
        '\n---\n**温馨提示',
        '\n\n**温馨提示**',
        '\n\n**参考链接**',
        '\n\n**相关链接**',
        # 分隔符后直接是引用链接的情况
        '\n\n---\n\n[',
    ]

    for marker in reference_markers:
        if marker in result:
            result = result.split(marker)[0]
            break

    # 如果还有 --- 后面跟着链接引用的内容，也要移除
    # 检查是否有类似 "[数字] [标题](链接)" 的引用格式
    lines = result.split('\n')
    filtered_lines = []
    in_reference_section = False

    for i, line in enumerate(lines):
        # 检测是否进入引用部分（以 [数字] 开头的链接行）
        if line.strip().startswith('[') and '] [' in line and '](' in line:
            in_reference_section = True
            break
        # 检测 --- 分隔符后紧跟引用的情况
        if line.strip() == '---' and i + 1 < len(lines):
            next_line = lines[i + 1]
            if next_line.strip().startswith('[') and '] [' in next_line:
                in_reference_section = True
                break
            # 或者 --- 后面是 ## 引用文件
            if '## 引用' in next_line or '## 参考' in next_line:
                in_reference_section = True
                break
        filtered_lines.append(line)

    if in_reference_section:
        result = '\n'.join(filtered_lines)

    return result.strip()


def chat_with_confirmation(question: str, max_rounds: int = 5) -> str:
    """
    执行多轮对话，自动处理确认
    只返回最终的结论（最后一轮不需要确认的回答）

    Args:
        question: 初始问题
        max_rounds: 最大对话轮数

    Returns:
        最终的AI回答（不包含确认过程和引用文件）
    """
    session_id = ""
    current_message = question
    last_content = ""

    for round_num in range(1, max_rounds + 1):
        print(f"\n=== 第 {round_num} 轮对话 ===")
        print(f"发送消息: {current_message[:50]}...")

        # 请求API
        api_response, session_id = request_api(current_message, session_id)
        print(f"Session ID: {session_id}")

        # 解析响应
        result = parse_response(api_response)
        content = result['full_content']
        last_content = content  # 保存最新一轮的回答

        print(f"AI回答长度: {len(content)} 字符")

        # 判断是否需要确认
        if is_confirmation_question(content):
            print("检测到需要确认的问题，自动回复'确认'...")
            current_message = "确认"
        else:
            print("已获得最终回答")
            break

    # 只返回最终的结论，并移除引用文件部分
    final_content = remove_references(last_content)
    print(f"\n最终结论长度: {len(final_content)} 字符")
    return final_content


def save_to_excel(question: str, answer: str, output_file: str = None):
    """
    将问题和回答保存到Excel表格

    Args:
        question: 用户问题
        answer: AI回答
        output_file: 输出文件名
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'AI回答结果'

    # 设置表头
    ws['A1'] = '问题'
    ws['B1'] = 'AI回答'

    # 设置列宽
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 100

    # 写入数据
    ws['A2'] = question
    ws['B2'] = answer

    # 设置自动换行
    ws['A2'].alignment = Alignment(wrap_text=True, vertical='top')
    ws['B2'].alignment = Alignment(wrap_text=True, vertical='top')

    # 生成文件名（带时间戳）
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"AI回答结果_{timestamp}.xlsx"

    # 保存文件
    wb.save(output_file)
    print(f"\n结果已保存到: {output_file}")


def read_questions_from_excel(filepath: str) -> list:
    """
    从Excel文件读取问题，从B2开始按行读取

    Args:
        filepath: Excel文件路径

    Returns:
        list: 问题列表，每个元素是 (row_number, question)
    """
    wb = load_workbook(filepath)
    ws = wb.active

    questions = []
    row = 2  # 从B2开始

    while True:
        cell_value = ws[f'B{row}'].value
        if cell_value is None or str(cell_value).strip() == '':
            break
        questions.append((row, str(cell_value).strip()))
        row += 1

    return questions


def save_results_to_excel(questions_with_answers: list, original_filepath: str, output_file: str = None):
    """
    保存结果到Excel，保留原格式

    Args:
        questions_with_answers: list of (row_number, question, answer)
        original_filepath: 原始文件路径
        output_file: 输出文件名
    """
    wb = load_workbook(original_filepath)
    ws = wb.active

    # 找到第一行空列的位置作为AI回答列
    max_col = ws.max_column
    answer_col = max_col + 1

    # 设置表头
    header_cell = ws.cell(row=1, column=answer_col)
    header_cell.value = 'AI回答'
    header_cell.font = Font(bold=True)

    # 写入回答
    for row_num, question, answer in questions_with_answers:
        cell = ws.cell(row=row_num, column=answer_col)
        cell.value = answer
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # 设置列宽
    ws.column_dimensions[ws.cell(row=1, column=answer_col).column_letter].width = 100

    # 生成文件名（带时间戳）
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"AI回答结果_{timestamp}.xlsx"

    # 保存文件
    wb.save(output_file)
    print(f"\n结果已保存到: {output_file}")


def process_excel_file(filepath: str):
    """
    处理Excel文件中的所有问题

    Args:
        filepath: Excel文件路径
    """
    print(f"读取文件: {filepath}")
    questions = read_questions_from_excel(filepath)

    if not questions:
        print("未找到问题（请确保从B2开始填写问题）")
        return

    print(f"共找到 {len(questions)} 个问题")

    results = []
    for i, (row_num, question) in enumerate(questions):
        print(f"\n{'='*50}")
        print(f"处理第 {i+1}/{len(questions)} 个问题 (行 {row_num})")
        print(f"问题: {question[:50]}...")

        try:
            answer = chat_with_confirmation(question)
            results.append((row_num, question, answer))
            print(f"回答长度: {len(answer)} 字符")
        except Exception as e:
            print(f"处理失败: {str(e)}")
            results.append((row_num, question, f"处理失败: {str(e)}"))

    # 保存结果
    save_results_to_excel(results, filepath)


def main():
    """主函数"""
    # 检查是否有命令行参数传入Excel文件
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
        process_excel_file(filepath)
    else:
        # 默认问题（单问题模式）
        question = "现行企业所得税法下，居民企业什么情况下取得的所得可能适用7.5%税率"

        print(f"原始问题: {question}")

        # 执行多轮对话
        final_answer = chat_with_confirmation(question)

        # 保存到Excel
        save_to_excel(question, final_answer)


if __name__ == "__main__":
    main()
